"""Backend code for buttons and query updates."""

from PyQt5 import QtWidgets
import openpyxl as pyxl
from openpyxl.utils import get_column_letter

from Apps.SalesResultsVisuals.LinkApplications import \
    SalesResultsAnalysisApplications
from Apps.ConnectDatabase import QSqlAuth, SQLiteAuth



class SalesResultsAnalysisBackend:
    def __init__(self, windowclass):
        """
        Attributes
        ---------------
        self.main : instance of Ui_sales_database
            Receives the GUI as instance.
        self.database_used : str
            Sets the database name that will be a parameter for
            multiple executions in methods.
        """
        self.main = windowclass
        self.database_used = "sales.db"

        self.__display_database_items()
        self.__connect_buttons()

    def __display_database_items(self):
        """Integrates database rows into PyQt5 SQL table."""
        query_line = self.read_query()
        with QSqlAuth(self.database_used) as datab:
            datab.qsql_show(self.main, "sales", query_line, "table_data")

    def __connect_buttons(self):
        """Connects buttons to correct method when pressed.
        Lambda used to avoid making additional Invoker methods."""
        self.open_app = SalesResultsAnalysisApplications()

        self.main.update_button.clicked.connect(
            lambda: self.open_app.sales_results_update())
        self.main.query_lineedit.returnPressed.connect(
            lambda: self.__update_query())
        self.main.extract_button.clicked.connect(
            lambda: self.__extract_query())
        self.main.visualize_button.clicked.connect(
            lambda: self.open_app.sales_visualize_param())

    def __update_query(self):
        """Updates table display based on query. If blank query,
        default query will be used."""
        query_line = self.read_query()
        if "local" in query_line:
            db_table = "local"
        else:
            db_table = "sales"

        with QSqlAuth(self.database_used) as datab:
            datab.qsql_show(self.main, db_table, query_line, "table_data")

    def __extract_query(self):
        """Extracts query based on user input on 'self.query_lineedit'"""
        extract_dir = QtWidgets.QFileDialog.getSaveFileName(filter="*.xlsx")

        if not extract_dir[0]:
            return

        extract_query = self.read_query()

        with SQLiteAuth(self.database_used) as datab:
            datab.sqlite_show(extract_query)
            extract_data = datab.data_extract
            database_columns = [
                headers[0] for headers in datab.data_headers
            ]  # Grabbing [0][0] and appending
        '''datab.data_headers returns a nested tuple in which the
        column name is on index [0]. Which is why we comprehend as above'''

        self.__prepare_workbook(
            extract_dir,
            extract_data,
            database_columns
        )

    def __prepare_workbook(self, filepath, output_data, output_columns):
        """Creates excel workbook with data based off query extract
        from method 'self.extract_query'"""
        extract_wb = pyxl.Workbook()
        extract_ws = extract_wb.active

        # Pasting column headers in row 1
        for index, data in enumerate(output_columns):
            extract_ws.cell(row=1, column=index + 1).value = data

        column_width = []
        for index, data in enumerate(output_data):
            for index2, data2 in enumerate(data):
                extract_ws.cell(
                    row=index + 2, column=index2 + 1
                ).value = data2  # +2 is because we start in row 2
                if len(column_width) > index2:
                    '''If the length of the list is smaller than the
                    index, then we still have to add column data'''
                    if len(str(data2)) > column_width[index2]:  #
                        column_width[index2] = len(str(data2))
                else:
                    column_width.append(len(str(data2)))

        for index, column_sizing in enumerate(column_width):
            extract_ws.column_dimensions[
                get_column_letter(index + 1)].width = column_sizing * 1.2

        extract_wb.save(filepath[0])

    def read_query(self):
        if not self.main.query_lineedit.text():  # Using default query
            return \
                "SELECT * FROM sales ORDER BY month, code, customer, bu1, bu2"
        else:
            return self.main.query_lineedit.text()