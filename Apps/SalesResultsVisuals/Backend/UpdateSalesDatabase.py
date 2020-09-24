"""Module with classes that handles the upload process based on
SAP data in an excel sheet.

Classes:
----------------------------------------------------------------------
SalesRFile - Holds the SAP excel workbook and base dictionaries
    for data columns

SalesRConfirmFormat - Holds composition of SalesRFile and checks:
    whether all data columns are in the excel sheet, whether number
    format is correct, whether month in file is same as combo box input

SalesRUpload - Holds composition of SalesRFile and controls the process
    of reading the correct data from the excel sheet and then writing
    it to the database

Functions:
-----------------------------------------------------------------------
finance_rounding - For each numerical data within the excel sheet,
    before writing to the database, it will enforce financial roundings
    for rounding up/down to 0 decimals
"""

import openpyxl as pyxl
import math
import re
from Apps.MessageBoxes import ErrorMessage, BasicMessage
from Apps.ConnectDatabase import SQLiteAuth

from Apps.ErrorClass import (
    IncorrectMonth, MonthNotFound, WrongNumberFormatting, HeaderMissing, NotConnectedError, NoFileSelected
)

class SalesRFile:
    """Creates a workbook/worksheet and dictionary attributes
    that contain headers as the key that are required to be in
    the worksheet.
    __________________________________________________________

    Class Attributes
    ---------------
    group : dict
    local : dict
    info : dict
        Group SAP data or local SAP data
        Key holds the headers that are required to be in the worksheet
        for reading/writing the correct data.
        Value at startup is "" and requires composition class
        'SalesRConfirmFormat' to match the header with the correct
        column index within the worksheet.
 """
    group = {
        "GROSS SALES": "",
        "NET SALES ex works": "",
        "Standard GM": "",
        "Standard CM1": "",
        "SALES VOLUME": ""
    }

    local = {
        "SALES VOLUME": "",
        "NET SALES ex works": "",
        "Actual GM": "",
        "CM1 variable local": "",
        "CM1 total/GP local": ""
    }

    info = {
        "Company code": "",
        "Rep. Customer": "",
        "B2B Level 1": "",
        "B2B Level 2": "",
        "Material (1-9)": ""
    }

    def __init__(self, dir, month):
        """Obtains file directory of excel file and month of results.

        Parameters
        -----------
        dir : str
            Absolute file path of excel file to be read/written from
        month : str
            Month of results in format '\d\d'

        Attributes
        -----------
        self.output_dir : str
            Holds the excel file directory
        self.sales_month : str
            Holds the sales result month in format '\d\d'
        self.wb, self.ws : openpyxl workbook/worksheet
            Opens the workbook with openpyxl, grabs the active sheet
        """
        self.output_dir = dir
        self.sales_month = month

        self.wb = pyxl.load_workbook(self.output_dir)
        self.ws = self.wb.active

class SalesRConfirmFormat:
    """Holds an instance of class 'SalesRFile'. Tests whether
    headers, date and numerical formattings are correct. Assigns
    correct column index to class attr dictionaries of the instance
    'SalesRFile'. If any formatting wrong, raises errors and stops.
    """
    def __init__(self, salesrclass):
        """Holds an instance of class 'SalesRFile'"""
        self.file = salesrclass

    def start_test(self):  # Default structure, can also methods call manually
        """Manager method for executing testing methods accordingly."""
        self.confirm_headers(self.file.group, 1)
        self.confirm_headers(self.file.info, 2)
        self.confirm_date()
        self.confirm_numbering()

    def confirm_headers(self, headers_dict, row):
        """Based on the Keys of param 'headers_dict', finds required
        headers in excel sheet. If found, writes column index of the
        header to Value. When testing row 2 in excel sheet, we add
        2 additional headers that are necessary for writing to db.

        Methods called
        -----------
        self._missing_headers : Checks whether Keys in dict all has
            a Value assigned. No Value assigned means header was not
            found and errors will be raised.
        """
        for column in range(1, self.file.ws.max_column + 1):
            header = self.file.ws.cell(row=row, column=column).value
            if header in headers_dict.keys():
                headers_dict[header] = column
        print(headers_dict)
        self._missing_headers(headers_dict, row)  # Checking for missing headers

        if row == 2: # Adds 2 additional headers for the None headers
            self.file.info["Company"] = \
                self.file.info["Company code"] + 1
            self.file.info["Material name"] = \
                self.file.info["Material (1-9)"] + 1

    def confirm_date(self):
        """Checks date in sheet whether it fits the date pattern and
        also whether it is the same date within the
        instance 'SalesRFile'.

        Raises
        -----------
        TypeError
            're' module can not match with NoneTypes. It is expected
            to have some blank cells in row 2, so we continue looping.

        IncorrectMonth
            Regex did not match month of instance 'SalesRFile'.
            This error gets raised to stop the program.

        MonthNotFound
            No month found in row 2, format is a mismatch. This error
            gets raised to stop the program."""
        month_in_file = re.compile(r"^\d\d\.\d\d\d\d$")
        for column in range(1, self.file.ws.max_column + 1):
            try:
                month_match = month_in_file.match(
                    self.file.ws.cell(row=2, column=column).value
                )
            except TypeError:
                month_match = None  # Blank cell

            if not month_match:
                continue
            elif self.file.sales_month in month_match[0]:
                return "Month in file matches month in combo box"
            else:
                ErrorMessage(
                    f"Month in file is displayed as {month_match[0]}, "
                    f"which does not match the month selected in the "
                    f"combo box ({self.file.sales_month})."
                )
                raise IncorrectMonth(
                    f"Month in file does not match month of combo box"
                )

        ErrorMessage(
            "No month was found in row 2. File formatting is a mismatch."
        )
        raise MonthNotFound(
            "Month cell not found in row 2. Check the excel formatting."
        )

    def confirm_numbering(self):
        """Checks whether a cell in row 3 contains '* 1.000'.
        Raises error if found, as this is the incorrect numbering.

        Raises
        -----------
        WrongNumberFormatting
            '* 1.000' found, thus numbering being wrong. Error gets
            raised to stop the program."""
        for column in range(1, self.file.ws.max_column + 1):
            if self.file.ws.cell(row=3, column=column).value == "* 1.000 ":
                ErrorMessage(
                    "Numbers are formatted as * 1.000. Please extract"
                    " the file from SAP with formatting as '1'"
                )
                raise WrongNumberFormatting("Formatted as * 1.000")

    def _missing_headers(self, headers_dict, row):
        """Dependency for method 'self.confirm_headers'. Checks
        whether all Keys have a value.

        Raises
        -----------
        HeaderMissing
            Not all Keys had a value. Error gets raised to stop the
            program."""
        if not all(headers_dict.values()):
            missing = [
                k for k in headers_dict.keys() if not headers_dict[k]
            ]
            ErrorMessage(
                f"Could not find {missing} as a column in row {row}. "
                f"Database updating canceled."
            )
            raise HeaderMissing(
                "Mandatory header not found. Either adjust mandatory headers"
                " in 'group'/'info' or fix the excel format"
            )


class SalesRUpload:
    """Holds an instance of class 'SalesRFile' and it is expected that
    the excel sheet has been succesfully tested by class
    'SalesRConfirmFormat. This class takes care of preparing the data
    and uploading to the database."""
    def __init__(self, salesrclass):
        """Holds instances of SalesRClass and database

        Attributes
        -----------
        self.file : class
            Instance of SalesRClass
        self.db : str
            Name of database for writing to
        self.database : class
            Instance of Facade class for db connection
        self.isConnected : Boolean
            True/False whether a db connection is active or not"""
        self.file = salesrclass
        self.db = "sales.db"
        self.database = SQLiteAuth(self.db)
        self.isConnected = False

    def open_db(self):
        """Opens database with SQLiteAuth instance."""
        self.database.sqlite_open()
        self.isConnected = True

    def close_db(self):
        """Closes database with SQLiteAuth instance."""
        self.database.sqlite_close()
        self.isConnected = False

    def confirm_upload_rows(self):
        """Creates a list of row index numbers that the 'start_upload'
        method will use for reading valid data. Invalid data are rows
        where 'Result' are present as a cell."""
        valid_rows = []
        for row in range(3, self.file.ws.max_row + 1):
            if self.file.ws.cell(row=row, column=3).value == "Result":
                continue
            elif self.file.ws.cell(row=row, column=4).value == "Result":
                continue
            elif self.file.ws.cell(row=row, column=5).value == "Result":
                continue
            elif self.file.ws.cell(row=row, column=6).value == "Result":
                continue
            elif self.file.ws.cell(row=row, column=1).value == "Overall Result":
                break  # End of table, No longer necessary to loop
            else:
                valid_rows.append(row)

        return valid_rows

    def _prepare_info(self, row):
        """Reads data from the row obtained through the parameter.
        Column will be grabbed from the headers dictionary.
        Returns data as list to method 'start_upload'."""
        code = self.file.ws.cell(
            row=row, column=self.file.info["Company code"]).value
        company = self.file.ws.cell(
            row=row, column=self.file.info["Company"]).value
        customer = self.file.ws.cell(
            row=row, column=self.file.info["Rep. Customer"]).value
        bu1 = self.file.ws.cell(
            row=row, column=self.file.info["B2B Level 1"]).value
        bu2 = self.file.ws.cell(
            row=row, column=self.file.info["B2B Level 2"]).value
        materialid = self.file.ws.cell(
            row=row, column=self.file.info["Material (1-9)"]).value
        material = self.file.ws.cell(
            row=row, column=self.file.info["Material name"]).value

        return [self.file.sales_month,
                code,
                company,
                customer,
                bu1,
                bu2,
                materialid,
                material
                ]

    def _prepare_group(self, row):
        """Reads data from the row obtained through the parameter.
        Column will be grabbed from the headers dictionary.
        Returns data as list to method 'start_upload'."""
        vol = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.group["SALES VOLUME"]).value)
        ns = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.group["NET SALES ex works"]).value)
        gs = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.group["GROSS SALES"]).value)
        gm = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.group["Standard GM"]).value)
        cm1 = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.group["Standard CM1"]).value)

        return [vol,
                ns,
                gs,
                gm,
                cm1
                ]

    def _prepare_local(self, row):
        """Reads data from the row obtained through the parameter.
        Column will be grabbed from the headers dictionary.
        Returns data as list to method 'start_upload'."""
        vol = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.local["SALES VOLUME"]).value)
        ns = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.local["NET SALES ex works"]).value)
        gs = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.local["CM1 variable local"]).value)
        gm = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.local["Actual GM"]).value)
        cm1 = finance_rounding(
            self.file.ws.cell(
                row=row, column=self.file.local["CM1 total/GP local"]).value)

        return [vol,
                ns,
                gs,
                gm,
                cm1
                ]

    def _upload_to_db(self, info_list, results_list):
        """Combines lists of the '_prepare' methods and turns it into
        one list, which will be used as the parameter for executing
        a query of data insertion to the database."""
        data = info_list + results_list

        if not self.isConnected:
            ErrorMessage("Code did not connect correctly to the database."
                         " Please check in 'UpdateSalesDatabase' in class "
                         "'SalesRUpload'")
            raise NotConnectedError("No connection with db.")

        self.database.execute_query(
            "CREATE TABLE IF NOT EXISTS sales ("
            "month TEXT, "
            "code TEXT, "
            "company TEXT, "
            "customer TEXT, "
            "bu1 TEXT, "
            "bu2 TEXT, "
            "materialid TEXT, "
            "material TEXT, "
            "vol INT, "
            "ns INT, "
            "gs INT, "
            "gm INT, "
            "cm1 INT)"
        )

        self.database.execute_query(
            "INSERT INTO sales VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            data
        )

    def start_upload(self):
        """Starts upload to database, uses data from dependency
        methods.

        Dependency methods
        -----------
        self.confirm_upload_rows - Obtains row indexes for valid data
        self.open_db - Creates connection to database
        self._prepare_info - Receives list of info data
        self._prepare_group - Receives list of sales results data
        self._upload_to_db - Combines data lists and writes to db
        self.close_db - Closes database

        Attributes
        -----------
        data_added, int - Counts total rows added to database
        valid_rows, list - Holds list of valid rows
        info_col, list - Holds list of info data
        results_col, list - Holds list of sales results data
        """
        data_added = 0

        valid_rows = self.confirm_upload_rows()

        self.open_db()

        if all(self.file.group.values()):
            for row in valid_rows:
                info_col = self._prepare_info(row)
                results_col = self._prepare_group(row)
                self._upload_to_db(info_col, results_col)
                data_added += 1
        else:
            for row in valid_rows:
                info_col = self._prepare_info(row)
                results_col = self._prepare_local(row)
                self._upload_to_db(info_col, results_col)
                data_added += 1

        self.close_db()

        BasicMessage(
            f"Writing completed. {data_added} new lines have been "
            f"added to the database."
        )

def finance_rounding(value):
    """Receives a float or int and returns a financially rounded
     number without decimals.

     TypeError turns NoneType cells into 0."""
    try:
        if value - math.floor(value) < 0.5:
            return math.floor(value)
        return math.ceil(value)
    except TypeError:
        return 0  # NoneType found, in results, change it to 0