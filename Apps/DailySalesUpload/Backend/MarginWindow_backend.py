from PyQt5.QtWidgets import QFileDialog

import openpyxl as pyxl
from dataclasses import dataclass

from Apps.ConnectDatabase import SQLiteAuth, QSqlAuth

from Apps.MessageBoxes import ErrorMessage, BasicMessage
from Apps.ErrorClass import UploadFailure

class MarginWindowBackend:
    def __init__(self, window):
        self.main = window
        self.database_used = "margins.db"

        self._connect_buttons()
        self._display_db()

    def _connect_buttons(self):
        self.main.extract_button.clicked.connect(
            lambda: "pass"
        )
        self.main.update_button.clicked.connect(
            lambda: self.update_database()
        )
        self.main.query_lineedit.returnPressed.connect(
            lambda: "pass"
        )

    def _display_db(self):
        query_line = self.read_query()
        with QSqlAuth(self.database_used) as datab:
            datab.qsql_show(self.main, "jmargin", query_line, "table_data")

    def update_database(self):
        dir = QFileDialog.getOpenFileName(filter="*.xlsx")

        if not dir[0]:
            return

        try:
            upload_data = MarginUpdater(dir[0])
            upload_data.upload_data()
        except UploadFailure:
            return

    def read_query(self):
        if not self.main.query_lineedit.text():
            return "SELECT * FROM jmargin ORDER BY product"
        else:
            product = self.main.query_lineedit.text()
            return f"SELECT * FROM jmargin WHERE product = '{product}'"


class MarginUpdater:
    def __init__(self, file_dir):
        self.margin_data = self._prepare_data(file_dir)
        self.database_used = "margins.db"

    def _prepare_data(self, dir):
        wb = pyxl.load_workbook(dir)
        ws = wb.active
        data = []

        try:
            for r in range(1, ws.max_row + 1):
                data.append(MarginData(
                    product=str(ws.cell(row=r, column=1).value),
                    price=int(ws.cell(row=r, column=2).value)
                ))
        except ValueError:
            ErrorMessage("In Column 'A' be sure to have dots/text in between "
                         "the numbers. In Column 'B be sure to only "
                         "have numbers. File match unsuccessful.")
            raise UploadFailure
        except TypeError:
            ErrorMessage("A row was found with no data, either remove"
                         " the row or add the missing data before trying"
                         " to upload.")
            raise UploadFailure

        return data

    def upload_data(self):
        updated_data = 0
        new_data = 0

        with SQLiteAuth(self.database_used) as datab:
            datab.execute_query(
                f"CREATE TABLE IF NOT EXISTS jmargin (product TEXT, price TEXT)"
            )
            datab.sqlite_show(
                f"SELECT product FROM jmargin")
            current_data = set(data[0] for data in datab.data_extract)

            for data_row in self.margin_data:
                upload_data = [data_row.product, data_row.price]
                if data_row.product in current_data:
                    # TODO - INSTEAD OF INTS, IT MAYBE IS OKAY TO INTEGRATE DOTS!! TAKE IT OUT IN THE END
                    # TODO - Figure out whether we have to convert price and product into a string
                    datab.execute_query(
                        f"UPDATE jmargin SET price = ? WHERE product = ?",
                        upload_data
                    )
                    updated_data += 1
                else:
                    # TODO - Figure out whether we have to convert price and product into a string
                    datab.execute_query(
                        f"INSERT INTO jmargin VALUES (?, ?)",
                        upload_data
                    )
                    new_data += 1

        BasicMessage(f"Succesfully uploaded!\n\n"
                     f"Product prices updated: {updated_data}\n"
                     f"Product prices newly added: {new_data}")

@dataclass(order=True)
class MarginData:
    product: int
    price: int












