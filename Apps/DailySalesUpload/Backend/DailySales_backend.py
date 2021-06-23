from PyQt5.QtWidgets import QFileDialog

import openpyxl as pyxl
import datetime
import os

from dataclasses import dataclass
from copy import deepcopy

from Apps.ConnectDatabase import SQLiteAuth
from Apps.DailySalesUpload.LinkApplications import DailySalesApplications
from Apps.MessageBoxes import ErrorMessage, BasicMessage
from Apps.ErrorClass import UploadFailure, UserCancel


class DailySalesBackend:
    def __init__(self, windowclass):
        """
        Attributes
        ---------------
        self.main : instance of Ui_sales_database
            Receives the GUI as instance.
        """
        self.main = windowclass
        self.prev_wb = None
        self.current_wb = None
        self.ooh_wb = None

        self.connect_buttons()

    def connect_buttons(self):
        self.main.upload_button.clicked.connect(
            lambda: self.get_dir(1)
        )
        self.main.upload_button_2.clicked.connect(
            lambda: self.get_dir(2)
        )
        self.main.upload_button_3.clicked.connect(
            lambda: self.get_dir(3)
        )
        self.main.save_button.clicked.connect(
            lambda: self.start_formatting()
        )

        self.new_window = DailySalesApplications()
        self.main.db_button.clicked.connect(
            lambda: self.new_window.show_margin_window()
        )

    def start_formatting(self):
        try:
            self.confirm_uploaded_files()
        except UploadFailure:
            return

        if self.main.comboBox.currentText() == "Yes":
            self._first_update()
        else:
            self._daily_update()

    def get_dir(self, button_number):
        dir = QFileDialog.getOpenFileName(filter="*.xlsx")

        if button_number == 1:
            if not dir[0]:
                self.main.uploaded_label.setText(
                    "Loading canceled, please select a file.")
                self.main.uploaded_label.setStyleSheet(
                    "background-color: rgb(255, 0, 0);")
            else:
                self.main.uploaded_label.setText(
                    f"Loaded: {os.path.basename(dir[0])}")
                self.main.uploaded_label.setStyleSheet(
                    "background-color: rgb(174, 217, 167);")
                self.prev_wb = dir[0]
        elif button_number == 2:
            if not dir[0]:
                self.main.uploaded_label_2.setText(
                    "Loading canceled, please select a file.")
                self.main.uploaded_label_2.setStyleSheet(
                    "background-color: rgb(255, 0, 0);")
            else:
                self.main.uploaded_label_2.setText(
                    f"Loaded: {os.path.basename(dir[0])}")
                self.main.uploaded_label_2.setStyleSheet(
                    "background-color: rgb(174, 217, 167);")
                self.current_wb = dir[0]
        else:
            if not dir[0]:
                self.main.uploaded_label_3.setText(
                    "Loading canceled, please select a file.")
                self.main.uploaded_label_3.setStyleSheet(
                    "background-color: rgb(255, 0, 0);")
            else:
                self.main.uploaded_label_3.setText(
                    f"Loaded: {os.path.basename(dir[0])}")
                self.main.uploaded_label_3.setStyleSheet(
                    "background-color: rgb(174, 217, 167);")
                self.ooh_wb = dir[0]

    def confirm_uploaded_files(self):
        if not self.prev_wb:
            ErrorMessage(
                "Please upload the former sales results file")
            raise UploadFailure(
                "Previous sales results file was not uploaded.")
        elif not self.current_wb and self.main.comboBox.currentText() == "No":
            ErrorMessage(
                "Please upload the new sales results file")
            raise UploadFailure(
                "New sales results file was not uploaded.")
        elif not self.ooh_wb and self.main.comboBox.currentText() == "No":
            ErrorMessage(
                "Please upload the main OOH file that should be located in "
                "the same directory of the previous sales upload file."
            )
            raise UploadFailure(
                "OOH file was not uploaded.")

    def _first_update(self):
        wb1 = SalesWorkbook(self.prev_wb)

        first_sales = DailyUpdater(wb1)

        try:
            first_sales.set_first_sales()
        except UserCancel:
            # TODO - If the list of data rows stays after cancel, we have to reset
            return

    def _daily_update(self):
        wb1 = SalesWorkbook(self.prev_wb)
        wb2 = SalesWorkbook(self.current_wb)
        ooh_wb = OOHWorkbook(self.ooh_wb)

        new_sales = DailyUpdater(last_wb=wb1, current_wb=wb2, ooh_wb=ooh_wb)

        try:
            new_sales.set_new_sales()
        except UserCancel:
            # TODO - If the list of data rows stays after cancel, we have to reset
            return

        # [print(data) for data in wb1.sales_data]


class SalesWorkbook:
    def __init__(self, dir):
        self.wb = pyxl.load_workbook(dir)
        self.ws = self._get_sales_ws(self.wb)
        self.sales_data_dates = self._get_sales_data_dates()

    def _get_sales_ws(self, workbook):
        for sheet in workbook.sheetnames:
            if workbook[sheet]['A1'].value == None:
                continue
            elif "売上" in workbook[sheet]['A1'].value:
                return workbook[sheet]
            else:
                continue

        ErrorMessage("Could not find the correct sheet where the "
                     "sales results are displayed. Be sure to have "
                     "the correct file uploaded.")
        raise UploadFailure("Incorrect sales results file uploaded.")

    def _get_sales_data_dates(self):
        sales_data = []

        for row_num in range(4, self.ws.max_row + 1):
            cust_id = self.ws.cell(row=row_num, column=2).value
            cust_name = self.ws.cell(row=row_num, column=1).value
            print(cust_name)
            product_id = self.ws.cell(row=row_num, column=7).value
            product_name = self.ws.cell(row=row_num, column=9).value
            delivery_date = self._adjust_month(
                self.ws.cell(row=row_num, column=4).value)
            kg = int(self.ws.cell(row=row_num, column=10).value)
            sales = int(self.ws.cell(row=row_num, column=12).value)

            if not sales_data:  # No data, so no concatinating possible yet
                sales_data.append(
                    SalesDataRow(cust_id,
                                 cust_name,
                                 product_id,
                                 product_name,
                                 kg,
                                 sales,
                                 invoice_date=delivery_date)
                )
                continue

            # Concatinating kg + sales on duplicate lines of customer-product
            for data in sales_data:
                existing_data = False
                if data.cust_id == cust_id and data.product_id == product_id:
                    if data.invoice_date == delivery_date:
                        data.kg += kg
                        data.sales += sales
                        existing_data = True
                        break

            if existing_data:
                continue

            # Non-existent data, so we add a new data line.
            sales_data.append(
                SalesDataRow(cust_id,
                             cust_name,
                             product_id,
                             product_name,
                             kg,
                             sales,
                             invoice_date=delivery_date)
            )

        return sales_data

    def _adjust_month(self, invoice_date):
        if invoice_date.month != datetime.datetime.now().month:
            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            day = 1

            new_date = datetime.datetime(year, month, day)

            while new_date.weekday() in [5, 6]:
                day += 1
                new_date = datetime.datetime(year, month, day)

            return new_date
        return invoice_date

class OOHWorkbook:
    def __init__(self, dir):
        self.wb = pyxl.load_workbook(dir)
        self.confirm_ws = self._get_confirm_ws()
        self.sales_data = self._get_sales_data()

    def _get_confirm_ws(self):
        for sheet in self.wb.sheetnames:
            if sheet == "Confirming Data":
                return self.wb[sheet]

        ErrorMessage("Could not find the correct sheet where the "
                     "OOH results are displayed. Be sure to have "
                     "the correct file uploaded.")
        raise UploadFailure("Incorrect OOH file uploaded.")

    def _get_sales_data(self):
        sales_data = []
        for row_num in range(2, self.confirm_ws.max_row + 1):
            cust_id = self.confirm_ws.cell(row=row_num, column=1).value
            cust_name = self.confirm_ws.cell(row=row_num, column=2).value
            product_id = self.confirm_ws.cell(row=row_num, column=3).value
            product_name = self.confirm_ws.cell(row=row_num, column=4).value
            kg = self.confirm_ws.cell(row=row_num, column=5).value
            sales = self.confirm_ws.cell(row=row_num, column=6).value
            margin_price = self.confirm_ws.cell(row=row_num, column=10).value
            cogs = self.confirm_ws.cell(row=row_num, column=7).value
            gross_margin = self.confirm_ws.cell(row=row_num, column=8).value
            profit_pct = self.confirm_ws.cell(row=row_num, column=9).value

            if isinstance(
                    self.confirm_ws.cell(
                        row=row_num, column=11
                    ).value, datetime.datetime):
                invoice_date = self.confirm_ws.cell(
                    row=row_num, column=11).value
            else:
                invoice_date = datetime.datetime.strptime(
                    self.confirm_ws.cell(row=row_num, column=11).value,
                    "%Y/%m/%d")

            sales_data.append(SalesDataRow(
                cust_id,
                cust_name,
                product_id,
                product_name,
                kg,
                sales,
                margin_price,
                cogs,
                gross_margin,
                profit_pct,
                invoice_date
            ))

        return sales_data


class DailyUpdater:
    def __init__(self, last_wb, current_wb=None, ooh_wb=None):
        self.margin_db = "margins.db"
        self.last_wb = last_wb
        self.current_wb = current_wb
        self.ooh_wb = ooh_wb

        self.ooh_to_sales = []
        self.new_sales_deviation = []
        self.new_sales_upload = []

    def set_new_sales(self):
        for latest_sales in self.current_wb.sales_data_dates:
            for old_sales in self.last_wb.sales_data_dates:
                is_updated_sales = self._detect_new_sales(
                    old_sales, latest_sales)
                if not is_updated_sales:
                    continue
                elif is_updated_sales != latest_sales:
                    self.new_sales_deviation.append(is_updated_sales)
                    break
                elif is_updated_sales == latest_sales:
                    break

            if not is_updated_sales:
                new_sales_row = deepcopy(latest_sales)
                self.new_sales_deviation.append(new_sales_row)

        self._collect_expired_ooh_data()
        self._combine_ooh_and_deviations()

        self._add_cogs_profit(self.new_sales_upload)
        self._add_cogs_profit(self.current_wb.sales_data_dates)

        sales_upload_dir = self._request_saving_dir()
        self._create_wb(self.new_sales_upload, sales_upload_dir)
        self._create_new_ooh_wb(sales_upload_dir)

    def set_first_sales(self):
        self._add_cogs_profit(self.last_wb.sales_data_dates)

        sales_upload_dir = self._request_saving_dir()
        self._create_new_ooh_wb(sales_upload_dir)
        self._create_wb(self.last_wb.sales_data_dates, sales_upload_dir)

    def _request_saving_dir(self):
        new_wb_dir = QFileDialog.getSaveFileName(filter="*.xlsx")

        if not new_wb_dir[0]:
            BasicMessage(
                "Upload canceled."
            )
            raise UserCancel

        return new_wb_dir[0]

    def _collect_expired_ooh_data(self):
        now = datetime.datetime.now().date()

        for expired_sales in self.ooh_wb.sales_data:
            if expired_sales.invoice_date.date() <= now:
                self.ooh_to_sales.append(expired_sales)

    def _combine_ooh_and_deviations(self):
        for sales_data in self.new_sales_deviation:
            cust_id = sales_data.cust_id
            prod_id = sales_data.product_id
            inv_date = sales_data.invoice_date.date()

            deviation_matched = False
            for ooh_data in self.ooh_to_sales:
                ooh_cust_id = ooh_data.cust_id
                ooh_prod_id = ooh_data.product_id
                ooh_inv_date = ooh_data.invoice_date.date()

                if cust_id == ooh_cust_id and prod_id == ooh_prod_id:
                    if inv_date == ooh_inv_date:
                        ooh_data.kg += sales_data.kg
                        ooh_data.sales += sales_data.sales
                        deviation_matched = True
                        continue

            if not deviation_matched:
                self.new_sales_upload.append(sales_data)

        for ooh_data in self.ooh_to_sales:
            self.new_sales_upload.append(ooh_data)

    def _detect_new_sales(self, old_sales, latest_sales):
        if old_sales.cust_id == latest_sales.cust_id:
            # Customer has sales, but need to check on same product.
            if old_sales.product_id == latest_sales.product_id:
                # Checking whether the sales on product date is the same
                if old_sales.invoice_date == latest_sales.invoice_date:
                    # Same product, but more sales could be made from before.
                    if not old_sales.sales == latest_sales.sales:
                        # New sales made on already-sold product
                        updated_sales = deepcopy(latest_sales)
                        updated_sales.kg = (
                                latest_sales.kg - old_sales.kg
                        )
                        updated_sales.sales = (
                                latest_sales.sales - old_sales.sales
                        )
                        return updated_sales
                    else:
                        # No changes in sales on already-sold product
                        return latest_sales
        return False

    def _add_cogs_profit(self, dataclasses_list):
        with SQLiteAuth(self.margin_db) as datab:
            datab.sqlite_show("SELECT * FROM jmargin")
            margin_data = datab.data_extract

        for data_row in dataclasses_list:
            for margin in margin_data:
                if margin[0] == data_row.product_id:
                    try:
                        data_row.margin_price = margin[1]
                        data_row.gross_margin = \
                            margin[1] * data_row.kg
                        data_row.cogs = \
                            data_row.sales - data_row.gross_margin
                        data_row.profit_pct = \
                            data_row.gross_margin / data_row.sales
                    except ZeroDivisionError:
                        data_row.profit_pct = 0

            if data_row.margin_price == "NOT FOUND" and data_row.cust_name in ["Nagano Sanyo", "Shoei", "Moriyama"]:
                data_row.margin_price = "TRADE BUSINESS"
                data_row.gross_margin = 0
                data_row.cogs = data_row.sales
                data_row.profit_pct = 0


        [print(m) for m in dataclasses_list]

    def _create_new_ooh_wb(self, sales_upload_dir):
        today = datetime.datetime.now()
        dir = os.path.join(
            os.path.dirname(sales_upload_dir),
            f"Orders on Hands.xlsx"
        )

        ooh_wb = pyxl.load_workbook("Excel Templates\\Orders on Hand.xlsx")
        ooh_confirm = ooh_wb["Confirming Data"]
        ooh_upload = ooh_wb["Upload Data"]

        if self.current_wb:
            sales_data_extract = self.current_wb.sales_data_dates
        else:
            sales_data_extract = self.last_wb.sales_data_dates

        starting_row = 2
        for sales in sales_data_extract:
            if sales.invoice_date >= datetime.datetime.now():
                print(f"{sales.invoice_date} VS {datetime.datetime.now()}")
                ooh_confirm.cell(
                    row=starting_row, column=1).value = sales.cust_id
                ooh_confirm.cell(
                    row=starting_row, column=2).value = sales.cust_name
                ooh_confirm.cell(
                    row=starting_row, column=3).value = sales.product_id
                ooh_confirm.cell(
                    row=starting_row, column=4).value = sales.product_name
                ooh_confirm.cell(
                    row=starting_row, column=5).value = sales.kg
                ooh_confirm.cell(
                    row=starting_row, column=6).value = sales.sales
                ooh_confirm.cell(
                    row=starting_row, column=7).value = sales.cogs
                ooh_confirm.cell(
                    row=starting_row, column=8).value = sales.gross_margin
                ooh_confirm.cell(
                    row=starting_row, column=9).value = sales.profit_pct
                ooh_confirm.cell(
                    row=starting_row, column=10).value = sales.margin_price
                ooh_confirm.cell(
                    row=starting_row, column=11).value = \
                    datetime.datetime.strftime(sales.invoice_date, "%Y/%m/%d")

                ooh_upload.cell(row=starting_row, column=1).value = \
                    datetime.datetime.strftime(sales.invoice_date, "%Y/%m/%d")
                ooh_upload.cell(row=starting_row, column=2).value = \
                    "3900"
                ooh_upload.cell(row=starting_row, column=3).value = \
                    sales.cust_id
                ooh_upload.cell(row=starting_row, column=4).value = \
                    sales.product_id.replace(".", "")
                ooh_upload.cell(row=starting_row, column=5).value = \
                    "3900"
                ooh_upload.cell(row=starting_row, column=6).value = \
                    sales.kg
                ooh_upload.cell(row=starting_row, column=7).value = \
                    sales.kg
                ooh_upload.cell(row=starting_row, column=8).value = \
                    sales.sales
                ooh_upload.cell(row=starting_row, column=9).value = \
                    "JPY"

                starting_row += 1

        ooh_wb.save(dir)

    def _create_wb(self, dataclasses_list, new_wb_dir):
        upload_wb = pyxl.load_workbook("Excel Templates\\Sales Upload.xlsx")
        confirm_ws = upload_wb["Confirming Data"]
        upload_ws = upload_wb["Upload Data"]

        starting_row = 2
        for sales in dataclasses_list:
            if sales.invoice_date <= datetime.datetime.now():
                confirm_ws.cell(
                    row=starting_row, column=1).value = sales.cust_id
                confirm_ws.cell(
                    row=starting_row, column=2).value = sales.cust_name
                confirm_ws.cell(
                    row=starting_row, column=3).value = sales.product_id
                confirm_ws.cell(
                    row=starting_row, column=4).value = sales.product_name
                confirm_ws.cell(
                    row=starting_row, column=5).value = sales.kg
                confirm_ws.cell(
                    row=starting_row, column=6).value = sales.sales
                confirm_ws.cell(
                    row=starting_row, column=7).value = sales.cogs
                confirm_ws.cell(
                    row=starting_row, column=8).value = sales.gross_margin
                confirm_ws.cell(
                    row=starting_row, column=9).value = sales.profit_pct
                confirm_ws.cell(
                    row=starting_row, column=10).value = sales.margin_price
                confirm_ws.cell(row=starting_row, column=11).value = \
                    datetime.datetime.strftime(sales.invoice_date, "%d.%m.%Y")

                upload_ws.cell(row=starting_row, column=1).value = \
                    "R"
                upload_ws.cell(row=starting_row, column=2).value = \
                    datetime.datetime.strftime(sales.invoice_date, "%d.%m.%Y")
                upload_ws.cell(row=starting_row, column=3).value = \
                    "3900"
                upload_ws.cell(row=starting_row, column=4).value = \
                    sales.cust_id  # remove the dots
                upload_ws.cell(row=starting_row, column=5).value = \
                    sales.product_id.replace(".", "")
                upload_ws.cell(row=starting_row, column=6).value = \
                    "VC"
                upload_ws.cell(row=starting_row, column=7).value = \
                    "JPY"
                upload_ws.cell(row=starting_row, column=8).value = \
                    sales.kg
                upload_ws.cell(row=starting_row, column=10).value = \
                    sales.sales
                upload_ws.cell(row=starting_row, column=18).value = \
                    sales.cogs
                upload_ws.cell(row=starting_row, column=19).value = \
                    sales.cogs

                starting_row += 1

        upload_wb.save(new_wb_dir)


@dataclass
class SalesDataRow:
    cust_id: str
    cust_name: str
    product_id: str
    product_name: str
    kg: int
    sales: int
    margin_price: margin_price = "NOT FOUND"
    cogs: cogs = "NONE CALCULATED"
    gross_margin: gross_margin = "NONE CALCULATED"
    profit_pct: profit_pct = "NONE CALCULATED"
    invoice_date: invoice_date = "NONE YET"
