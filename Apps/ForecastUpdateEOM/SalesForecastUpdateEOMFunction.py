import pyautogui
import re
import os
import pprint
import copy
import sys
import openpyxl as pyxl
import datetime
from pathlib import Path
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

# Data structures:
sales_data = {} #Our data structure
storage_delivery_trade = {'delivery_storage': 0,
                          'trade_volume': 0,
                          'trade_sales': 0,
                          'trade_profit': 0} # for storing the total net sales of all storage data and trade business

trade_customers = ['Moriyama', 'Nagano Sanyo', 'Shoei'] # Add here additional trade customers if they will ever come

trade_business_error = []

def UploadEOMResults(sales_file, forecast_file, reporting_month):
    # Section 1 - Getting the path and file based on user input - Sales Result file and Sales FC file

    sales_wb = pyxl.load_workbook(sales_file)
    for sheet in sales_wb.sheetnames:
        if sales_wb[sheet]['A1'].value == None:
            continue
        elif "売上" in sales_wb[sheet]['A1'].value:
            sales_sheet = sales_wb[sheet]
            break

    for cell in range(4, sales_sheet.max_row + 1):
        customer = sales_sheet.cell(row=cell, column=1).value
        productID = sales_sheet.cell(row=cell, column=7).value

        # Sometimes there are human errors where product name will have a space at the end
        # Also sometimes formatting goes wrong (wrapping text in cell), which disrupts the dictionary value
        # e.g. product_name is 'Berries', but due to formatting error, dictionary picked up 'Berries_xnxc_\n'
        # Therefore it will not match the next if statement, because the product_name is not in dictionary

        product = sales_sheet.cell(row=cell, column=9).value
        if '_x000D_\n' in product:
            print(product)
            product = product.replace('_x000D_\n', '')
            print(product)

        volume = sales_sheet.cell(row=cell, column=10).value
        sales = sales_sheet.cell(row=cell, column=12).value
        # If for some reason a cell is empty in the volume/sales column, we make it 0 to prevent errors when converting
        # to int() or float() in the section below
        if not volume:
            volume = 0
        if not sales:
            sales = 0

        # Now start filling up the dictionary with /getdefault

        sales_data.setdefault(customer, {})
        sales_data[customer].setdefault(productID, {})
        sales_data[customer][productID].setdefault(product, {'volume': 0, 'sales': 0, 'profit': 0})

        sales_data[customer][productID][product]['volume'] += round(float(volume), 2)
        sales_data[customer][productID][product]['sales'] += int(sales)

        if sales_sheet.cell(row=cell, column=13).value == None:  # if no cogs in file, then we just put 0 as profit
            sales_data[customer][productID][product]['profit'] += 0
        elif type(sales_sheet.cell(row=cell, column=13).value) == type(
                str()):  # this has 'NO COGS' message, profit = sales
            sales_data[customer][productID][product]['profit'] += int(sales)
        else:  # has cogs, we calculate the profit by sales - cogs
            sales_data[customer][productID][product]['profit'] += int(sales) - sales_sheet.cell(row=cell,
                                                                                                column=13).value

    pprint.pprint(sales_data)

    forecast_wb = pyxl.load_workbook(forecast_file)

    forecast_sheet = forecast_wb['2020 All']
    sales_summary_sheet = forecast_wb["Summary RESULTS"]  # For updating the trade business results in the summary table

    unmatched_sales = copy.deepcopy(sales_data)  # we want to keep the original sales_data dict as it is.
    # for this dict, we will take out all succesfull links and only keep the unmatched sales so that we can check manually
    # and input it into the sales forecast ourself

    # Section 2 - Formatting the forecast to its EOM state and so that the forecast is ready for use for next month

    for cell in range(1, forecast_sheet.max_column):
        if re.search(r'SALES.*VOLUME', str(forecast_sheet.cell(row=1, column=cell).value)):
            if forecast_sheet.cell(row=2, column=cell).value == reporting_month:
                print(f'{cell} - VOLUME')
                EOM_Vol = cell  # Column for pasting volume data
        elif re.search(r'Net.*Sales', str(forecast_sheet.cell(row=1, column=cell).value)):
            if forecast_sheet.cell(row=2, column=cell).value == reporting_month:
                print(f'{cell} - NS')
                EOM_Sales = cell  # Column for pasting sales data
        elif re.search(r'JM', str(forecast_sheet.cell(row=1, column=cell).value)):
            if forecast_sheet.cell(row=2, column=cell).value == reporting_month:
                print(f'{cell} - JM')
                EOM_JM = cell  # Column for pasting JM data

    # Change MTD header to next month, Change YTD header to current reporting month
    next_reporting_month = datetime.datetime.strptime(reporting_month, '%b')
    next_reporting_month = next_reporting_month.replace(month=next_reporting_month.month + 1)
    next_reporting_month = next_reporting_month.strftime('%b')  # Converts the current month + 1, so the month after

    MTD_Naming = forecast_sheet.cell(row=2, column=19).value
    forecast_sheet.cell(row=2, column=19).value = re.sub('\w\w\w$', next_reporting_month,
                                                         MTD_Naming)  # changing MTD Header
                                                        # Using .sub for only changing the MMM in string

    # Change YTD header to current completed reporting month
    YTD_Naming = forecast_sheet.cell(row=2, column=17).value
    forecast_sheet.cell(row=2, column=17).value = re.sub('\d\d\d\d\w\w\w$',
                                                         f'{datetime.datetime.now().year}{reporting_month}',
                                                         YTD_Naming)  # Changing YTD Header

    # Grab the columns of EOM, change the formula in the cells so that it also grabs this one, do this for all rows
    # In value of cell, we it must be '=SUM(T' as starting, if it is e.g. '=SUM(Q', then this is no relevant
    greenFill = PatternFill(start_color='0099CC00', end_color='0099CC00', fill_type='solid')  # Filling cell to green

    # Deleting all MTD Data before writing new in just to be sure that no relevant data is in the MTD
    for cell in range(3, forecast_sheet.max_row + 1):
        if forecast_sheet.cell(row=cell, column=10).value != "SQLCOPY":
            continue
        forecast_sheet.cell(row=cell, column=19).value = ""
        forecast_sheet.cell(row=cell, column=37).value = ""
        forecast_sheet.cell(row=cell, column=55).value = ""

    for cell in range(1, forecast_sheet.max_row + 1):
        if not forecast_sheet.cell(row=cell, column=17).value:  # Continues over empty cells
            continue
        if forecast_sheet.cell(row=cell, column=17).value.startswith('=SUM(T'):
            forecast_sheet.cell(row=cell,
                                column=17).value = f'=SUM(T{cell}:{get_column_letter(EOM_Vol)}{cell})'  # '=SUM(T3:Y3)
            forecast_sheet.cell(row=cell,
                                column=33).value = f'=SUM(AL{cell}:{get_column_letter(EOM_Sales)}{cell})'  # sales
            forecast_sheet.cell(row=cell,
                                column=51).value = f'=SUM(BD{cell}:{get_column_letter(EOM_JM)}{cell})'  # sales

            # Colorize the row of the current reporting month
            forecast_sheet.cell(row=cell, column=EOM_Vol).fill = greenFill
            forecast_sheet.cell(row=cell, column=EOM_Vol).value = ''

            forecast_sheet.cell(row=cell, column=EOM_Sales).fill = greenFill
            forecast_sheet.cell(row=cell, column=EOM_Sales).value = ''

            forecast_sheet.cell(row=cell, column=EOM_JM).fill = greenFill
            forecast_sheet.cell(row=cell, column=EOM_JM).value = ''

    for cell in range(3, forecast_sheet.max_row):
        customer_cell = forecast_sheet.cell(row=cell, column=11).value
        prod_ID = forecast_sheet.cell(row=cell, column=14).value
        product_name = forecast_sheet.cell(row=cell, column=15).value

        if customer_cell in sales_data:
            print(f'{customer_cell} data found. Checking Product ID')

            if prod_ID in sales_data[customer_cell]:
                print(f'{customer_cell} matches {prod_ID}. Checking Product Name')

                for product_key in sales_data[customer_cell][prod_ID].keys():
                    print(f'{product_name} matches {prod_ID} and {customer_cell}. Giving value')
                    print(f"Sales: {sales_data[customer_cell][prod_ID][product_key]['sales']}")
                    forecast_sheet.cell(row=cell, column=EOM_Vol).value = sales_data[customer_cell][prod_ID][product_key][
                        'volume']
                    forecast_sheet.cell(row=cell, column=EOM_Sales).value = sales_data[customer_cell][prod_ID][product_key][
                        'sales']
                    forecast_sheet.cell(row=cell, column=EOM_JM).value = sales_data[customer_cell][prod_ID][product_key][
                        'profit']

                    if product_name != product_key:
                        trade_business_error.append(f"Please change {customer_cell} {prod_ID} {product_name} > {product_key}")

                # if product_name in sales_data[customer_cell][prod_ID]:
                #     print(f'{product_name} matches {prod_ID} and {customer_cell}. Giving value')
                #     print(f"Sales: {sales_data[customer_cell][prod_ID][product_name]['sales']}")
                #
                #     forecast_sheet.cell(row=cell, column=EOM_Vol).value = sales_data[customer_cell][prod_ID][product_name][
                #         'volume']
                #     forecast_sheet.cell(row=cell, column=EOM_Sales).value = sales_data[customer_cell][prod_ID][product_name][
                #         'sales']
                #     forecast_sheet.cell(row=cell, column=EOM_JM).value = sales_data[customer_cell][prod_ID][product_name][
                #         'profit']

                    try:
                        del unmatched_sales[customer_cell][prod_ID]  # delete sales line, because it has been matched

                        if unmatched_sales[customer_cell] == {}:
                            del unmatched_sales[
                                customer_cell]  # If no more values/products are left, we delete the whole customer
                            continue
                    except KeyError:
                        pyautogui.alert(
                            f"{unmatched_sales[customer_cell][prod_ID]} seems to have a duplicate in the FC.\n"
                            f"Please delete the duplicate row and try again.")
                        sys.exit()

    unmatched_sales2 = copy.deepcopy(unmatched_sales)

    # From unmatched_data, we need to check whether unmatched_data[customer][prodID] is Delivery'/'Storage fee ID
    print("Unmatched data loaded...")
    for cust in unmatched_sales:  # we will only need the total delivery/storage fee, so we concatinate all of them into a new dic
        for prodID in unmatched_sales[cust]:
            if prodID in ['00001', '00003', '00008', '00001-A']:
                for prodN in unmatched_sales[cust][prodID]:
                    storage_delivery_trade['delivery_storage'] += unmatched_sales[cust][prodID][prodN]['sales']

                    del unmatched_sales2[cust][prodID]  # deleting line, because it has been matched

                    if unmatched_sales2[cust] == {}:
                        del unmatched_sales2[cust]
    print("Deliver and storage data loaded...")

    unmatched_sales3 = copy.deepcopy(unmatched_sales2)

    for cust in unmatched_sales2:
        if cust in trade_customers:  # trade_customers has been defined at start of script. For if we need to add more
            for prodID in unmatched_sales2[cust]:
                for prodN in unmatched_sales2[cust][prodID]:  # Only concatinating the int
                    storage_delivery_trade['trade_volume'] += unmatched_sales2[cust][prodID][prodN]['volume']
                    storage_delivery_trade['trade_sales'] += unmatched_sales2[cust][prodID][prodN]['sales']
                    storage_delivery_trade['trade_profit'] += unmatched_sales2[cust][prodID][prodN]['profit']  # see next if

                    if unmatched_sales2[cust][prodID][prodN]['profit'] != 0:  # we will still copy to xlsx, but will warn user
                        trade_business_error.append(f'{sales_data[cust][prodID]} - Profit should be 0 as trading business,\n' \
                                               f'Keeping this line in the unmatched sales tab for checking,\n' \
                                               f'Please confirm whether the COGS is OKAY. Profit has been added to JM')

                        del unmatched_sales3[cust][prodID][prodN]['sales']
                        del unmatched_sales3[cust][prodID][prodN]['volume']
                    else:

                        del unmatched_sales3[cust][prodID]

                    if unmatched_sales3[cust] == {}:
                        del unmatched_sales3[cust]

    for cell in range(1, forecast_sheet.max_row + 1):
        if forecast_sheet.cell(row=cell, column=15).value == 'Trade Business (Nagano/Haruna/Shoei)':
            forecast_sheet.cell(row=cell, column=19).value = ''
            forecast_sheet.cell(row=cell, column=37).value = ''
            forecast_sheet.cell(row=cell, column=55).value = ''

            forecast_sheet.cell(row=cell, column=EOM_Vol).value = storage_delivery_trade['trade_volume']
            print('Pasted trade volume.')
            forecast_sheet.cell(row=cell, column=EOM_Sales).value = storage_delivery_trade['trade_sales'] + \
                                                             storage_delivery_trade[
                                                                 'delivery_storage']  # combining for Net sales
            print('Pasted trade sales.')
            forecast_sheet.cell(row=cell, column=EOM_JM).value = storage_delivery_trade['trade_profit'] + \
                                                             storage_delivery_trade['delivery_storage']
            # DeliveryStorage has no Cogs, so all profit
            print('Pasted trade profit.')
            break

    # As a last step, we will go into the summary sheet and update the trade business results
    trade_storage_delivery = 0  # There might be multiple similar cells, but we only want the first 3
    volume_trade = []
    # Finding the rows for Volume, sales and JM
    for cell in range(1, sales_summary_sheet.max_row + 1):
        if sales_summary_sheet.cell(row=cell, column=2).value == "Trade Business/Storage/Shipping":
            print(f"Found Volume Trade/Storage/Shipping in row {cell}")
            volume_trade.append(cell)
            trade_storage_delivery += 1
        if trade_storage_delivery == 3:
            break

    # Finding the column for Volume, Sales and JM
    volume_trade_column = None
    cell_content = ["Volume", "Net Sales", "Japan Margin"]
    for content in cell_content:
        for cell in range(1, sales_summary_sheet.max_row + 1):
            if sales_summary_sheet.cell(row=cell, column=2).value == content:
                # Found the row where we have to look for the correct month
                for cell2 in range(2, sales_summary_sheet.max_column + 1):
                    if sales_summary_sheet.cell(row=cell, column=cell2).value == reporting_month:
                        volume_trade_column = cell2

    # Pasting the trade volume, trade sales and trade profit in the summary
    sales_summary_sheet.cell(row=volume_trade[0], column=volume_trade_column).value = storage_delivery_trade['trade_volume']
    sales_summary_sheet.cell(row=volume_trade[1], column=volume_trade_column).value = storage_delivery_trade['trade_sales'] + \
                                                             storage_delivery_trade[
                                                                 'delivery_storage']
    sales_summary_sheet.cell(row=volume_trade[2], column=volume_trade_column).value = storage_delivery_trade['trade_profit'] + \
                                                             storage_delivery_trade['delivery_storage']








    # For some reason it makes the width of the columns too big...
    # forecast_sheet.column_dimensions.group(start="T", end="AE", outline_level=0)
    # forecast_sheet.column_dimensions.group(start="AL", end="AW", outline_level=0)
    # forecast_sheet.column_dimensions.group(start="BD", end="BO", outline_level=0)
    # forecast_sheet.column_dimensions.group(start="T", end=get_column_letter(EOM_Vol), hidden=True)
    # forecast_sheet.column_dimensions.group(start="AL", end=get_column_letter(EOM_Sales), hidden=True)
    # forecast_sheet.column_dimensions.group(start="BD", end=get_column_letter(EOM_JM), hidden=True)



    return [sales_data, unmatched_sales3, storage_delivery_trade, trade_business_error, forecast_wb]