import os

import openpyxl as pyxl
from openpyxl.utils import get_column_letter

from PyQt5 import QtWidgets

from Apps.ErrorClass import HeaderMissing
from Apps.ErrorClass import WrongMonthFormatting
from Apps.ErrorClass import UploadFailure
from Apps.ErrorClass import NoDataFound

from Apps.ConnectDatabase import QSqlAuth
from Apps.ConnectDatabase import SQLiteAuth

from Apps.LTOAnalysis.Backend.ltoAnalysisUpdater import LTOUpdate


class LTOAnalysisBackend:
    def __init__(self, windowclass):
        self.main = windowclass
        self.database_used = "LTO.db"

        self.update_query()
        self.__connect_buttons()

    def __connect_buttons(self):
        self.main.update_button.clicked.connect(
            lambda: self.update_lto_clicked())
        self.main.query_lineedit.returnPressed.connect(
            lambda: self.update_query())
        self.main.extract_button.clicked.connect(
            lambda: self.extract_query())

    def update_lto_clicked(self):
        lto_dir = QtWidgets.QFileDialog.getOpenFileName(filter="*.xlsx")

        if not lto_dir[0]:
            self.main.update_label.setText(
                "Update file not inserted.")
            self.main.update_label.setStyleSheet(
                "background-color: rgb(255, 0, 0);")
            return

        lto_dir = os.path.abspath(lto_dir[0])
        self.main.update_label.setText(
            f"{os.path.basename(lto_dir)} - Updating database, please wait..."
        )

        update_lto = LTOUpdate(lto_dir)
        try:
            update_lto.start_formatting()
            update_lto.start_uploading()
        except (HeaderMissing, WrongMonthFormatting,
                UploadFailure, NoDataFound):
            return

        self.main.update_label.setText("Database updated!")

        self.update_query()

    def update_query(self):
        query_line = self.read_query()

        with QSqlAuth(self.database_used) as datab:
            datab.qsql_show(self.main, "lto", query_line, "table_data")

    def extract_query(self):
        extract_dir = QtWidgets.QFileDialog.getSaveFileName(
            filter="*.xlsx")

        if not extract_dir[0]:
            return

        extract_query = self.read_query()

        with SQLiteAuth(self.database_used) as datab:
            datab.sqlite_show(extract_query)
            headers = [header[0] for header in datab.data_headers]
            extract_data = datab.data_extract

        extract_wb = ExtractedQueryWb()
        extract_wb.add_headers(headers)
        extract_wb.add_query_data(extract_data, col_width=True)
        extract_wb.save(extract_dir[0])

    def read_query(self):
        if not self.main.query_lineedit.text():
            return "SELECT * from lto ORDER BY launch"
        else:
            return self.main.query_lineedit.text()

class ExtractedQueryWb:
    def __init__(self):
        self.wb = pyxl.Workbook()
        self.ws = self.wb.active
        self.column_width = []

    def add_headers(self, headers):
        for index, header in enumerate(headers):
            self.ws.cell(row=1, column=index + 1).value = header

    def add_query_data(self, extracted_data, col_width=False):
        for index, data_items in enumerate(extracted_data):
            for index2, data_item in enumerate(data_items):
                self.ws.cell(
                    row=index + 2, column=index2 + 1).value = data_item
                if col_width:
                    self._measure_col_width(index2, data_item)

        if col_width:
            self._fix_column_sizing()

    def _measure_col_width(self, index, data):
        # If the length of list is smaller than index, then we stil lhave to add column data
        if len(self.column_width) > index:
            if len(str(data)) > self.column_width[index]:
                self.column_width[index] = len(str(data))
        else:
            self.column_width.append(len(str(data)))

    def _fix_column_sizing(self):
        for index, column_sizing in enumerate(self.column_width):
            self.ws.column_dimensions[
                get_column_letter(index + 1)].width = column_sizing * 1.2

    def reset_column_width(self):
        self.column_width = []

    def save(self, dir):
        self.wb.save(dir)