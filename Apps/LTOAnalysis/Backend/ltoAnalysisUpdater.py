import openpyxl as pyxl
from openpyxl.utils import get_column_letter
from Apps.MessageBoxes import ErrorMessage, BasicMessage
from Apps.ErrorClass import HeaderMissing, UploadFailure, WrongMonthFormatting, NotFormatted, NoDataFound
from Apps.ConnectDatabase import SQLiteAuth
import datetime as dt
import re

class LTOUpdate:
    def __init__(self, wb_dir):
        self.format = LTOFormatFile(wb_dir)
        self.upload = LTOFileUpload(self.format)

    def start_formatting(self):
        self.format.find_starting_rows()
        self.format.define_headers()
        self.format.check_missing_headers()
        self.format.set_month_format()
        self.format.hide_unnecessary_columns()

    def start_uploading(self):
        self.upload.db.sqlite_open()

        self.upload.confirm_existing_data()
        self.upload.write_to_db()
        self.upload.remove_old_data()
        self.upload.display_results()

        self.upload.db.sqlite_close()


class LTOFormatFile:
    original_headers = {
        "Opportunity ID": "",
        "Opportunity Name": "",
        "Customer": "",
        "Customer Responsibility": "",
        "Opportunity Stage": "",
        "Launch Date": "",
        "EUR": "",
        "VOL": ""  # This will always be location of EUR + 1
    }

    def __init__(self, wb_directory):
        self.dir = wb_directory
        self.wb = pyxl.load_workbook(self.dir)
        self.ws = self.wb.active

        self.header_row = None
        self.first_data_row = None
        self.isCorrectFormat = False

    def find_starting_rows(self):
        self._find_header_row()
        self._find_first_data_row()

    def _find_header_row(self):
        """Defines the correct row of the headers."""
        self.header_row = 1
        last_row = self.ws.max_row
        while True:
            if self.ws.cell(
                    row=self.header_row, column=1
            ).value == "Opportunity ID":
                break  # Found correct row, break loop
            elif self.header_row >= last_row:
                ErrorMessage(
                    f"Could not find 'Opportunity ID' as a column."
                    f"Database updating canceled."
                )
                raise HeaderMissing  # Excel file is wrong
            self.header_row += 1  # Not found yet, loop next row

    def _find_first_data_row(self):
        """Defines the starting row of the first data line."""
        find_id = re.compile(r"^\d+$")
        for row in range(self.header_row, self.ws.max_row + 1):
            if find_id.match(self.ws.cell(row=row, column=1).value):
                self.first_data_row = row
                break

        if not self.first_data_row:
            ErrorMessage(
                "Could not find starting point of data lines. "
                "A correct Opportunity ID was not found in Column A."
            )
            raise NoDataFound()

    def define_headers(self):
        for col in range(1, self.ws.max_row + 1):
            header = self.ws.cell(row=self.header_row, column=col).value
            if header in self.original_headers.keys():
                self.original_headers[header] = col

    def check_missing_headers(self):
        try:
            self.original_headers["VOL"] = self.original_headers["EUR"] + 1
        except TypeError:
            ErrorMessage(
                f"Could not find 'EUR' as a column."
                f"Database updating canceled."
            )
            raise HeaderMissing()

        if not all(self.original_headers.values()):
            missing = [
                k for k in self.original_headers.keys()
                if not self.original_headers[k]
            ]

            ErrorMessage(
                f"Could not find '{missing}' as a column."
                f"Database updating canceled."
            )
            raise HeaderMissing("Not all headers found when"
                                "reading file format.")

        self.isCorrectFormat = True

    def set_month_format(self):
        if not self.isCorrectFormat:
            raise NotFormatted(
                "File was not checked for missing headers yet."
            )

        for row in range(self.first_data_row, self.ws.max_row + 1):
            try:
                unformatted_month = self.ws.cell(
                    row=row,
                    column=self.original_headers["Launch Date"]).value
                formatting_month = dt.datetime.strptime(
                    unformatted_month, "%d.%m.%Y")
                self.ws.cell(
                    row=row,
                    column=self.original_headers["Launch Date"]).value = \
                    formatting_month.strftime("%Y/%m/%d")
            except:
                ErrorMessage(
                    "The date at header 'Launch Date' could not be formatted "
                    "correctly. If the format is not DD.MM.YYYY, then "
                    "either the LTO file is wrong or the code must be "
                    "updated so that it reads the format correctly."
                )
                raise WrongMonthFormatting()
            finally:
                self.wb.save(f"{self.dir} 1")

    def hide_unnecessary_columns(self):
        if not self.isCorrectFormat:
            raise NotFormatted(
                "File was not checked for missing headers yet."
            )

        total_cols = set(col for col in range(1, self.ws.max_column + 1))
        data_cols = set(col for col in self.original_headers.values())
        hide_cols = data_cols.symmetric_difference(total_cols)

        for col in hide_cols:
            hide_col = get_column_letter(col)
            if not self.ws.column_dimensions[hide_col].hidden:
                self.ws.column_dimensions[hide_col].hidden = True

        self.wb.save(f"{self.dir} 1")


class LTOFileUpload:
    def __init__(self, ltofile_class):
        self.file = ltofile_class
        self.db = SQLiteAuth("lto.db")

        self.new_adds = 0
        self.existing_adds = 0
        self.closed_adds = 0

        self.existing_data = []
        self.uploaded_data = []

    def confirm_existing_data(self):
        if self.existing_data:
            raise UploadFailure(
                "List 'self.existing_data' already has data appended, "
                "meaning the data from the specific sheet has already "
                "been uploaded."
            )

        self.file.check_missing_headers()

        self.db.sqlite_show("SELECT id from lto")
        for opp_id in range(len(self.db.data_extract)):
            self.existing_data.append(self.db.data_extract[opp_id][0])

    def write_to_db(self):
        """
        Attributes
        -----------
        starting_row : int
            starting row point, which is 1 row under the header.
        """
        starting_row = self.file.first_data_row
        for cell in range(starting_row, self.file.ws.max_row + 1):
            id = int(self.file.ws.cell(
                row=cell,
                column=self.file.original_headers[
                    "Opportunity ID"]).value)
            name = self.file.ws.cell(
                row=cell,
                column=self.file.original_headers[
                    "Opportunity Name"]).value
            cus = self.file.ws.cell(
                row=cell,
                column=self.file.original_headers[
                    "Customer"]).value
            kam = self.file.ws.cell(
                row=cell,
                column=self.file.original_headers[
                    "Customer Responsibility"]).value
            stage = self.file.ws.cell(
                row=cell,
                column=self.file.original_headers[
                    "Opportunity Stage"]).value
            launch = self.file.ws.cell(
                row=cell,
                column=self.file.original_headers[
                    "Launch Date"]).value
            sales = self.file.ws.cell(
                row=cell,
                column=self.file.original_headers[
                    "EUR"]).value
            vol = self.file.ws.cell(
                row=cell,
                column=self.file.original_headers[
                    "VOL"]).value
            status = "None"
            comment = "None"

            self.uploaded_data.append(id)
            if id in self.existing_data:
                self.existing_adds += 1
                data_update = [name, cus, kam, stage, launch, sales, vol, id]
                self.db.execute_query("UPDATE lto SET "
                                      "name = ?, "
                                      "cus = ?, "
                                      "kam = ?, "
                                      "stage = ?, "
                                      "launch = ?, "
                                      "sales = ?, "
                                      "vol = ? "
                                      "WHERE id = ?",
                                      data_update)
            else:
                self.new_adds += 1
                data_new = [
                    id, name, cus, kam, stage,
                    launch, sales, vol, status, comment]
                self.db.execute_query("INSERT INTO lto VALUES "
                                      "(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                                      data_new)

    def remove_old_data(self):
        for remove_data in self.existing_data:
            if remove_data not in self.uploaded_data:
                self.closed_adds += 1
                print(
                    f"ID {remove_data} is old data, removing from database..."
                )
                self.db.execute_query(
                    "DELETE FROM lto WHERE id = ?",
                    (str(remove_data),))

    def display_results(self):
        self.db.sqlite_show("SELECT * FROM lto")
        BasicMessage(f"New added: {self.new_adds}\n"
                   f"Existing adjusted: {self.existing_adds}\n"
                   f"Removed old data: {self.closed_adds}\n\n"
                   f"Total in Table: {len(self.db.data_extract)}")
