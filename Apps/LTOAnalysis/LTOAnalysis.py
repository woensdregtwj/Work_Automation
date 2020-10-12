# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'ltoSQLTableEnlarged.ui'
#
# Created by: PyQt5 UI code generator 5.13.2
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
import os
from UpdateLTODatabase import *
from openpyxl.utils import get_column_letter

from Apps.LTOAnalysis.Backend.ltoAnalysisUpdater import LTOUpdate
from Apps.ErrorClass import HeaderMissing, WrongMonthFormatting, UploadFailure, NoDataFound
from Apps.ConnectDatabase import QSqlAuth, SQLiteAuth

from Apps.LTOAnalysis.Backend.LTOAnalysis_backend import LTOAnalysisBackend


class Ui_lto_database(object):
    def setupUi(self, lto_database):
        lto_database.setObjectName("lto_database")
        lto_database.resize(1600, 805)
        lto_database.setMinimumSize(QtCore.QSize(1600, 800))
        lto_database.setMaximumSize(QtCore.QSize(16777215, 16777215))
        lto_database.setStyleSheet("background-color: rgb(0, 40, 72);")
        self.centralwidget = QtWidgets.QWidget(lto_database)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.setObjectName("verticalLayout")
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setMinimumSize(QtCore.QSize(1601, 80))
        self.frame.setStyleSheet("background-color: rgb(0, 35, 72);")
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.logo_label = QtWidgets.QLabel(self.frame)
        self.logo_label.setGeometry(QtCore.QRect(0, 0, 101, 81))
        self.logo_label.setText("")
        self.logo_label.setPixmap(QtGui.QPixmap("Images/DoehlerLogo.png"))
        self.logo_label.setScaledContents(True)
        self.logo_label.setObjectName("logo_label")
        self.app_title = QtWidgets.QLabel(self.frame)
        self.app_title.setGeometry(QtCore.QRect(100, 0, 1491, 81))
        self.app_title.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("MS Reference Sans Serif")
        font.setPixelSize(32)
        font.setBold(False)
        font.setWeight(50)
        font.setKerning(True)
        self.app_title.setFont(font)
        self.app_title.setStyleSheet("background-color: rgb(0, 35, 72);\n"
                                     "color: rgb(255, 255, 255);")
        self.app_title.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.app_title.setFrameShadow(QtWidgets.QFrame.Plain)
        self.app_title.setMidLineWidth(0)
        self.app_title.setTextFormat(QtCore.Qt.PlainText)
        self.app_title.setAlignment(QtCore.Qt.AlignCenter)
        self.app_title.setWordWrap(False)
        self.app_title.setObjectName("app_title")
        self.verticalLayout.addWidget(self.frame)
        self.actions_frame = QtWidgets.QFrame(self.centralwidget)
        self.actions_frame.setMinimumSize(QtCore.QSize(1601, 71))
        self.actions_frame.setStyleSheet("background-color: rgb(0, 40, 72);")
        self.actions_frame.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.actions_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.actions_frame.setObjectName("actions_frame")
        self.query_lineedit = QtWidgets.QLineEdit(self.actions_frame)
        self.query_lineedit.setGeometry(QtCore.QRect(10, 30, 711, 31))
        font = QtGui.QFont()
        font.setPixelSize(15)
        self.query_lineedit.setFont(font)
        self.query_lineedit.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.query_lineedit.setObjectName("query_lineedit")
        self.query_label = QtWidgets.QLabel(self.actions_frame)
        self.query_label.setGeometry(QtCore.QRect(10, 0, 261, 21))
        font = QtGui.QFont()
        font.setPixelSize(20)
        font.setBold(True)
        font.setUnderline(False)
        font.setWeight(75)
        font.setKerning(True)
        self.query_label.setFont(font)
        self.query_label.setStyleSheet("color: rgb(255, 255, 255);")
        self.query_label.setObjectName("query_label")

        self.visualize_button = QtWidgets.QPushButton(self.actions_frame)
        self.visualize_button.setGeometry(QtCore.QRect(850, 10, 181, 51))
        font = QtGui.QFont()
        font.setPixelSize(15)
        self.visualize_button.setFont(font)
        self.visualize_button.setStyleSheet("color: rgb(255, 255, 255);\n"
                                            "background-color: rgb(0, 111, 196);")
        self.visualize_button.setObjectName("visualize_button")

        self.extract_button = QtWidgets.QPushButton(self.actions_frame)
        self.extract_button.setGeometry(QtCore.QRect(740, 10, 100, 51))
        font = QtGui.QFont()
        font.setPixelSize(10)
        self.extract_button.setFont(font)
        self.extract_button.setStyleSheet("color: rgb(255, 255, 255);\n"
                                            "background-color: rgb(0, 111, 196);")
        self.extract_button.setObjectName("visualize_button")

        self.update_button = QtWidgets.QPushButton(self.actions_frame)
        self.update_button.setGeometry(QtCore.QRect(1040, 10, 181, 51))
        font = QtGui.QFont()
        font.setPixelSize(15)
        self.update_button.setFont(font)
        self.update_button.setStyleSheet("color: rgb(255, 255, 255);\n"
                                         "background-color: rgb(0, 111, 196);")
        self.update_button.setObjectName("update_button")
        self.update_label = QtWidgets.QLabel(self.actions_frame)
        self.update_label.setEnabled(True)
        self.update_label.setGeometry(QtCore.QRect(1230, 20, 361, 31))
        font = QtGui.QFont()
        font.setPixelSize(10)
        self.update_label.setFont(font)
        self.update_label.setStyleSheet("color: rgb(255, 255, 255);\n"
                                        "border-color: rgb(255, 255, 255);\n"
                                        "background-color: rgb(0, 85, 127);")
        self.update_label.setObjectName("update_label")
        self.tableload_label = QtWidgets.QLabel(self.actions_frame)
        self.tableload_label.setGeometry(QtCore.QRect(570, 0, 171, 21))
        font = QtGui.QFont()
        font.setPixelSize(20)
        font.setBold(True)
        font.setUnderline(False)
        font.setWeight(75)
        font.setKerning(True)
        self.tableload_label.setFont(font)
        self.tableload_label.setStyleSheet("color: rgb(255, 255, 255);")
        self.tableload_label.setObjectName("tableload_label")
        self.verticalLayout.addWidget(self.actions_frame)
        self.table_data = QtWidgets.QTableView()
        self.table_data.setMinimumSize(QtCore.QSize(1581, 611))
        self.table_data.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.table_data.setObjectName("table_data")
        font = QtGui.QFont()
        font.setPixelSize(14)
        self.table_data.setFont(font)
        self.verticalLayout.addWidget(self.table_data)
        lto_database.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(lto_database)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1600, 18))
        self.menubar.setObjectName("menubar")
        lto_database.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(lto_database)
        self.statusbar.setObjectName("statusbar")
        lto_database.setStatusBar(self.statusbar)

        self.retranslateUi(lto_database)
        QtCore.QMetaObject.connectSlotsByName(lto_database)

        back = LTOAnalysisBackend(self)

        # self.database_used = "LTO.db"
        #
        # # self.query = QSqlQuery(db=db)
        # self.update_query()

        # self.model = QSqlTableModel(db=db)
        # self.model.setTable("lto")
        # self.model.setEditStrategy(QSqlTableModel.OnRowChange)
        #
        # self.query = QSqlQuery(db=db)
        # self.query.prepare("SELECT * FROM lto ORDER BY launch") # setSort does not working, have to do through Query
        # self.query.exec_()
        # self.model.setQuery(self.query)
        #
        # self.table_data.setModel(self.model)
        # self.table_data.resizeColumnsToContents()

        # self.update_button.clicked.connect(self.update_lto_clicked)
        # self.query_lineedit.returnPressed.connect(self.update_query)
        # self.extract_button.clicked.connect(self.extract_query)

    def retranslateUi(self, lto_database):
        _translate = QtCore.QCoreApplication.translate
        lto_database.setWindowTitle(_translate("lto_database", "MainWindow"))
        self.app_title.setText(_translate("lto_database", "LTO Analysis - Doehler Japan"))
        self.query_label.setText(_translate("lto_database", "Query Box (SQLite syntax)"))
        self.visualize_button.setText(_translate("lto_database", "Visualize displayed data"))
        self.extract_button.setText(_translate("lto_database,", "Extract shown data"))
        self.update_button.setText(_translate("lto_database", "Update database"))
        self.update_label.setText(
            _translate("lto_database", "To update  - Please select excel file originating from \'WE LEAD ANALYTICS\'"))
        self.tableload_label.setText(_translate("lto_database", "lto table loaded"))
#
#     def update_lto_clicked(self):
#         lto_dir = QtWidgets.QFileDialog.getOpenFileName(filter="*.xlsx")
#
#         if not lto_dir[0]:
#             self.update_label.setText(
#                 "Update file not inserted.")
#             self.update_label.setStyleSheet(
#                 "background-color: rgb(255, 0, 0);")
#             return
#
#         lto_dir = os.path.abspath(lto_dir[0])
#         self.update_label.setText(
#             f"{os.path.basename(lto_dir)} - Updating database, please wait..."
#         )
#
#         update_lto = LTOUpdate(lto_dir)
#         try:
#             update_lto.start_formatting()
#             update_lto.start_uploading()
#         except (HeaderMissing, WrongMonthFormatting,
#                 UploadFailure, NoDataFound):
#             return
#
#         self.update_label.setText("Database updated!")
#
#         self.update_query()
#
#     def update_query(self):
#         query_line = self.read_query()
#
#         with QSqlAuth(self.database_used) as datab:
#             datab.qsql_show(self, "lto", query_line, "table_data")
#
#     def extract_query(self):
#         self.extract_dir = QtWidgets.QFileDialog.getSaveFileName(
#             filter="*.xlsx")
#
#         if not self.extract_dir[0]:
#             return
#
#         extract_query = self.read_query()
#
#         with SQLiteAuth(self.database_used) as datab:
#             datab.sqlite_show(extract_query)
#             headers = [header[0] for header in datab.data_headers]
#             extract_data = datab.data_extract
#
#         extract_wb = ExtractedQueryWb()
#         extract_wb.add_headers(headers)
#         extract_wb.add_query_data(extract_data, col_width=True)
#         extract_wb.save(self.extract_dir[0])
#
#     def read_query(self):
#         if not self.query_lineedit.text():
#             return "SELECT * from lto ORDER BY launch"
#         else:
#             return self.query_lineedit.text()
#
#
# class ExtractedQueryWb:
#     def __init__(self):
#         self.wb = pyxl.Workbook()
#         self.ws = self.wb.active
#         self.column_width = []
#
#     def add_headers(self, headers):
#         for index, header in enumerate(headers):
#             self.ws.cell(row=1, column=index + 1).value = header
#
#     def add_query_data(self, extracted_data, col_width=False):
#         for index, data_items in enumerate(extracted_data):
#             for index2, data_item in enumerate(data_items):
#                 self.ws.cell(
#                     row=index + 2, column=index2 + 1).value = data_item
#                 if col_width:
#                     self._measure_col_width(index2, data_item)
#
#         if col_width:
#             self._fix_column_sizing()
#
#     def _measure_col_width(self, index, data):
#         # If the length of list is smaller than index, then we stil lhave to add column data
#         if len(self.column_width) > index:
#             if len(str(data)) > self.column_width[index]:
#                 self.column_width[index] = len(str(data))
#         else:
#             self.column_width.append(len(str(data)))
#
#     def _fix_column_sizing(self):
#         for index, column_sizing in enumerate(self.column_width):
#             self.ws.column_dimensions[
#                 get_column_letter(index + 1)].width = column_sizing * 1.2
#
#     def reset_column_width(self):
#         self.column_width = []
#
#     def save(self, dir):
#         self.wb.save(dir)
#
#
#

if __name__ == "__main__":
    import sys

    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")
    lto_database = QtWidgets.QMainWindow()
    ui = Ui_lto_database()
    ui.setupUi(lto_database)
    lto_database.show()
    sys.exit(app.exec_())
