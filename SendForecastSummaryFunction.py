import openpyxl as pyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import *
import re
import datetime
import logging

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# We want to copy ONLY the numbers from the forecast file sheet "Summary RESULTS".
# Load workbook as data only being true, otherwise it will copy the values

def create_summary(forecast_file, overhead, fx_rate):
    global wb, ws, new_wb, new_ws
    wb = pyxl.load_workbook(forecast_file, data_only=True)
    ws = wb["Summary RESULTS"]

    new_wb = pyxl.load_workbook("Excel Templates\\Sales forecast summary.xlsx")
    new_ws = new_wb.active

    if fx_rate:
        global rate
        rate = int(fx_rate)
        logging.debug(f"{rate} found, going to convert...")
    else:
        rate = None

    # Copying data
    copy_paste(5, 8)  # Copying (A) Volume lines
    copy_paste(10, 13)  # Copying (B) Volume lines
    copy_paste(15, 18, step=2)  # Copying (C) and "Adjust" Volume lines

    copy_paste(23, 26)  # Copying (A) Net Sales
    copy_paste(28, 31)  # Copying (B) Net Sales
    copy_paste(33, 36, step=2)  # Copying (C) and "Adjust" Net Sales

    copy_paste(41, 44)  # Copying (A) JM
    copy_paste(46, 49)  # Copying (B) JM
    copy_paste(51, 54, step=2)  # Copying (C) JM

    # Pasting the overhead costs
    for column in range(2, 14):
        new_ws.cell(row=55, column=column).value = overhead

    return new_wb

def copy_paste(r1, r2, step=1):
    for row in range(r1, r2, step):
        copy_row = []
        # Copying 1 row of data
        for column in range(3, 15):
            copy_row.append(ws.cell(row=row, column=column).value)

        logging.debug(f"Appended row {row}")
            # Pasting 1 row of data
        index = 0
        for column_paste in range(2, 14):
            if rate and r1 not in [5, 10, 15]:
                logging.debug(f"Rate conversion starting")
                new_ws.cell(row=row, column=column_paste).value = copy_row[index] / rate
            else:
                logging.debug(f"No rate conversion")
                new_ws.cell(row=row, column=column_paste).value = copy_row[index]
            index += 1






