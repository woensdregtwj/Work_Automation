import openpyxl as pyxl
import datetime
import sqlite3

def gx_format(worksheet):
    test_ws = worksheet

    # These are in A1, A2, A3, A4, A5, A6. Should always be fixed
    fixed_items = ["株式会社ドーラジャパン", "勘定科目: 142", "補助科目", "細目", "部門", "承認レベル"]

    # Testing whether this is the correct format.
    for index, item in enumerate(fixed_items):
        if not test_ws.cell(row=index + 1, column=1).value:  # If cell is empty, format incorrect
            failed_test = f"Nothing in cell 'A{index + 1}, while '{item}' should be written. Test failed."
            return failed_test
        elif isinstance(test_ws.cell(row=index + 1, column=1).value, int):  # If int in cell, format incorrect
            failed_test = f"Integer found in 'A{index + 1}'. There should be no integers. Test failed."
            return failed_test
        elif item not in test_ws.cell(row=index + 1, column=1).value:  # If fixed item not in cell, format incorrect
            failed_test = f"Format seems to be incorrect. We could not match '{item}' at 'A{index + 1}'"
            return failed_test
    return True

def convert_aging_list(workbook, month, type):
    wb = pyxl.load_workbook(workbook)
    ws = wb.active  # Original file will only have 1 sheet, so you can use ws = wb.active

    upload_wb = pyxl.load_workbook("Excel Templates\\Receivables_Upload.xlsx")
    upload_ws = upload_wb.active

    conn = sqlite3.connect("Databases\\payment_terms.db")
    c = conn.cursor()
    # Database columns are: customer, id, term

    reporting_month = datetime.datetime.strptime(month, "%b").month  # Change this into being an argument for the
    # function.
    if type == "Monthly":  # Reporting type has to be converted to a number for the SAP format
        reporting_type = 12
    else:
        reporting_type = 1

    payment_data = {}  # Data structure will be {customerid: Account Receivable}
    delete_payment_data = []  # Duplicates with "-1" in its cusid will be concatenated to its original and deleted
    unmatched_customers = {}  # Data structure as {customername: customerid}

    is_correct_format = gx_format(ws)
    if isinstance(is_correct_format, str):
        print("We failed the format test. Returning error text to UI.")
        print(is_correct_format)
        return [is_correct_format]  # This will be put in an error popup box so that the user can check the format.

    # Defining the starting row with data. According to the GX file output.
    # It starts at row with string "Accounts Receivable" on column B.
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "142" and ws.cell(row=row, column=2).value == "Accounts Receivable":
            starting_row = row
            break

    # From starting_row, we loop over each customerid and AR amount and add this to our dictionary.
    for row in range(starting_row, ws.max_row + 1):
        if not ws.cell(row=row, column=6).value and not ws.cell(row=row + 1, column=6).value:
            print("Reached the end, exiting loop")
            break
        payment_data.setdefault(ws.cell(row=row, column=6).value, ws.cell(row=row, column=10).value)

    # Now we look for keys with "-" in its name, we have to delete these and concatinate the value to the key without "-"
    for dup_data in payment_data.keys():
        if "-" in dup_data:
            original_data = dup_data.split("-")
            payment_data[original_data[0]] += payment_data[dup_data]

            delete_payment_data.append(dup_data)

    # We delete the keys with "-" from our dictionary, as we concatinated the amount to its non "-" key
    for dup_data in delete_payment_data:
        print(f"Deleting {dup_data} from our data")
        del payment_data[dup_data]

    # Now that we have the data for being written to the upload template, we can start pasting the dictionary.
    paste_in_row = 2
    for key, value in payment_data.items():
        paste_row = [datetime.datetime.now().year, reporting_month, key, 3900, "JPY", "", reporting_type, "Legal", 0, 0, 0,
                     0, value, value, 0, 0, 0, 0, 0, 0, 0, 0]  # This is 1 row to be pasted. There are a lot of default
        # data, so change here if any values have to be changed. The empty "" is where we will paste the payment term
        print(f"Pasting in row {paste_in_row} - {paste_row}")

        for index, paste_value in enumerate(paste_row):
            upload_ws.cell(row=paste_in_row, column=index + 1).value = paste_value
        paste_in_row += 1

    # Here we call the payment_terms.db and match the correct payment terms with its SAP customer number code;
    # If the SAP code is not in the database, then assign it as 60
    # Database columns are: customer, id, term

    for term_cell in range(2, len(payment_data) + 2):
        match_term = upload_ws.cell(row=term_cell, column=3).value

        c.execute("SELECT term FROM terms WHERE id = ?", (match_term,))
        is_matched = c.fetchall()

        try:
            if not is_matched:
                print("Not matched, payment term will be 60 days. Writing to unmatched dict for reference")

                for cus_row in range(starting_row, ws.max_row + 1):  # This is from the uploaded wb, not new wb
                    if ws.cell(row=cus_row, column=6).value == match_term:
                        match_cus_term = ws.cell(row=cus_row, column=5).value

                unmatched_customers.setdefault(match_cus_term, match_term)  # {customer name: customer id}
                upload_ws.cell(row=term_cell, column=6).value = 60
            else:
                print(f"{match_term} has a payment term of {is_matched}")
                upload_ws.cell(row=term_cell, column=6).value = str(is_matched).strip("'[(,)]'")
        except:
            print(f"{is_matched} - Is this a duplicate?")
            return [f"{is_matched} shows up 2 times in the database, please remove and try again."]

    return [upload_wb, unmatched_customers]


# return workbook
# return how many duplicates were deleted from the sql database.
# if we ever had to delete a duplicate, then it means that it the database should be checked for mistakes!!!!