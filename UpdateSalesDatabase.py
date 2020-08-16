import openpyxl as pyxl
from openpyxl.utils import get_column_letter
import sqlite3
import math


# TODO - Create a test for checking whether the template is okay
def UpdateSalesResultsTest(workbook, month):
    wb = pyxl.load_workbook(workbook)
    ws = wb.active

    headers = ["GROSS SALES", "NET SALES ex works", "Standard GM", "Standard CM1", "SALES VOLUME"]
    headers2 = ["Company code", None, "Rep. Customer", "B2B Level 1", "B2B Level 2", "Material (1-9)", None]
    headers_location = {}

    # Testing empty cells at headers
    for c in range(1, 8):
        column_letter = get_column_letter(c)
        if ws.cell(row=1, column=c).value:
            print(f"TEST FAILED 1 - cell is supposed to be empty at '{column_letter}1'")
            return f"Header failed - cell is supposed to be empty at '{column_letter}1'"
    print("Succesfully passed test 1 - Empty cells are empty in A1:G1")

    # Testing Column headers data
    for c in range(8, 13):
        header = ws.cell(row=1, column=c).value
        report_date = ws.cell(row=2, column=c).value
        #num_format = ws.cell(row=3, column=c).value
        column_letter = get_column_letter(c)

        if header not in headers:
            print(f"TEST FAILED 2 - Unexpected value found in '{column_letter}1'")
            return f"Header 2 failed - Unexpected value '{header}' found in '{column_letter}1'." \
                   f"Headers should only be: \nGROSS SALES, NET SALES ex works, Standard GM, Standard CM1, SALES VOLUME"

        if month not in report_date:
            print(f"TEST FAILED 2 - Wrong month header has been detected '{column_letter}2'")
            return f"Header 2 failed- Wrong month header has been detected in '{column_letter}2'.\n" \
                   f"Month in file shows '{report_date}', while month header should be '{month}.YYYY'.\n\n" \
                   f"Please select the correct month from the upload window."

        # if num_format != "* 1.000 ":
        #     print(f"TEST FAILED 2 - number format is not in *1.000 '{column_letter}3'")
        #     return f"Header 2 failed - number format is not in *1.000 at '{column_letter}3'."

        # Headers are OKAY, we save the header locations for writing to the database
        headers_location.setdefault(header, c)
        headers.remove(header)  # Removing header from the list to avoid duplicates being found in the excel sheet
    print("Succesfully passed test 2 - all correct headers found.\n"
          "Dictionary will be passed on for writing to database.")

    # Testing Column headers data 2
    for index, value in enumerate(headers2):
        column_letter = get_column_letter(index+1)
        if value != ws.cell(row=2, column=index+1).value:
            print(f"TEST FAILED 3 - Wrong header title at row 3 column {index+1}")
            return f"Header 3 failed - Wrong header title at '{column_letter}2'. Please be sure to have the order of " \
                   f"the headers the same as shown as in the example image."
    print("Succesfully passed test 3. Testing has been finalized!")

    return headers_location  # Dictionary with the column numbers will be used for grabbing data for the database

def finance_rounding(value):
    if value - math.floor(value) < 0.5:
        return math.floor(value)
    return math.ceil(value)

# TODO - Create a new database if not exist.
# Columns: Month, Company code, Company Name, Customer, B2B 1, B2B 2, Material, Material name, Volume, NS, GS, GM, CM1
def create_sales_db(workbook, month, column_dict):
    connect = sqlite3.connect("Databases\\sales.db")
    c = connect.cursor()
    print("Connected")

    c.execute("CREATE TABLE IF NOT EXISTS sales (month TEXT, code TEXT, company TEXT, customer TEXT, bu1 TEXT, bu2 TEXT, "
              "materialid TEXT, material TEXT, vol INT, ns INT, gs INT, gm INT, cm1 INT)")
    print("Created table if not existed, along with the columns")

# TODO - Start looping over the rows and copy the designated cell to the correct column in the database
# If column C, D, E, F == "Result", then skip over this. This is a summed value that we want to exclude.
# If Column A == "Overall Result", then skip and end. This is the end of the table in excel.
    wb = pyxl.load_workbook(workbook)
    ws = wb.active
    print("reading into workbook and worksheet")

    lines_added = 0
    for cell in range(3, ws.max_row + 1):
        # Skipping over the sub-sum "Result" rows in the sheet. They add no value to the db
        if ws.cell(row=cell, column=3).value == "Result" or ws.cell(row=cell, column=4).value == "Result":
            continue
        if ws.cell(row=cell, column=5).value == "Result" or ws.cell(row=cell, column=6).value == "Result":
            continue

        # Defining the end of the table and breaking the loop so that it does not loop over empty cells afterwards
        if ws.cell(row=cell, column=1).value == "Overall Result":
            print("End of table reached. Exiting loop.")
            break

        # If there are any NoneTypes in the integer area, we change them to 0 to avoid TypeError
        for data in range(8, ws.max_column + 1):
            if not ws.cell(row=cell, column=data).value:
                ws.cell(row=cell, column=data).value = 0

        # Starting the ACTUAL writing to the database

        # Setting the values that will be written to the db
        # month = month
        code = ws.cell(row=cell, column=1).value
        company = ws.cell(row=cell, column=2).value
        customer = ws.cell(row=cell, column=3).value
        bu1 = ws.cell(row=cell, column=4).value
        bu2 = ws.cell(row=cell, column=5).value
        materialid = ws.cell(row=cell, column=6).value
        material = ws.cell(row=cell, column=7).value

        vol = finance_rounding(ws.cell(row=cell, column=column_dict["SALES VOLUME"]).value)
        ns = finance_rounding(ws.cell(row=cell, column=column_dict["NET SALES ex works"]).value)
        gs = finance_rounding(ws.cell(row=cell, column=column_dict["GROSS SALES"]).value)
        gm = finance_rounding(ws.cell(row=cell, column=column_dict["Standard GM"]).value)
        cm1 = finance_rounding(ws.cell(row=cell, column=column_dict["Standard CM1"]).value)

        # Writing the data to the database
        data_write = [month, code, company, customer, bu1, bu2, materialid, material, vol, ns, gs, gm, cm1]
        print(data_write)

        c.execute("INSERT INTO sales VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", data_write)
        print("Inserted")
        connect.commit()

        lines_added += 1

    print(f"Writing completed. {lines_added} new lines have been added")

# TODO - Close database and return something to the user to confirm that the data has succesfully been added
    c.close()
    connect.close()
    print("Database closed")

    return lines_added

# TODO - Based on the combobox, check whether there already is data in the database of that specific month.
# If it is, return back to the user and ask for input: Is this okay, should we overwrite? Or cancel?
# DO THIS IN THE CLASS OF THE WINDOW BEFORE EXECUTING ALL FUNCTIONS ABOVE!!!!
