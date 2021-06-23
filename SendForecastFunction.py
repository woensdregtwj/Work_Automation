import openpyxl as pyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import *
import re
import datetime
import logging


logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# This function is supposed to create a list of all the necessary ints from the forecast.
# Each list will resemble the row within the MAIN forecast sheet and will be copied to a template.
# The template is saved in this environ.

# List structure will look like this:

# [Customer Name, Customer Responsibility, Material ID, Material Name, MTD, VolJan, VolFeb, volMar, volApr -... etc.

# Our def will have 3 arguments:
# Forecast_file > The main forecast file from where we will get our data from for copying
# manager > The Gui can specify which manager's data to grab. If manager is "All", will create 1 file per manager
# month > Necessary so that we can hide the previous months, as these are irrelevant and make reading more difficult
# Month is also necessary for creating the YTD formula and also for renaming the columns

def main_forecast_file(forecast_file, manager, current_month, fx_rate):
    main_wb = pyxl.load_workbook(forecast_file)
    main_ws = main_wb['All']

    new_wb = pyxl.load_workbook("Excel Templates\\Sales forecast.xlsx")
    new_ws = new_wb.active

    kam = manager
    month = current_month

    if fx_rate:
        global rate
        rate = int(fx_rate)
        logging.debug(f"{rate} found, going to convert...")
    else:
        rate = None

    for cell in range(3, main_ws.max_row + 1):
        main_data = []
        mtd_data = []
        pricing = []
        volume = []
        sales = []
        jm = []
        # forecast_data_row = [[main_data], [mtd_data], [pricing], [volume], [sales], [jm]]

        if main_ws.cell(row=cell, column=10).value == "SQLCOPY" and main_ws.cell(row=cell, column=12).value == kam:
            # Appending main data
            main_data.append(main_ws.cell(row=cell, column=11).value)  # Customer Name
            main_data.append(main_ws.cell(row=cell, column=12).value)  # KAM
            main_data.append(main_ws.cell(row=cell, column=14).value)  # Material ID
            main_data.append(main_ws.cell(row=cell, column=15).value)  # Material Name
            logging.debug(f'Appended main data: {main_data} ')

            if rate:
                # Appending MTD data
                mtd_data.append(main_ws.cell(row=cell, column=19).value)  # MTD Volume

                if not main_ws.cell(row=cell, column=37).value:
                    mtd_data.append(main_ws.cell(row=cell, column=37).value)  # MTD Net Sales
                    mtd_data.append(main_ws.cell(row=cell, column=55).value)  # MTD JM, usually 0 if not EOM data
                else:
                    mtd_data.append(main_ws.cell(row=cell, column=37).value / rate)  # MTD Net Sales
                    mtd_data.append(main_ws.cell(row=cell, column=55).value / rate)  # MTD JM, usually 0 if not EOM data
                logging.debug(f'Appended mtd data: {mtd_data} ')

                # Appending NS Price and GM Price
                if type(main_ws.cell(row=cell, column=35).value) in [type(1), type(0.1)]:
                    pricing.append(main_ws.cell(row=cell, column=35).value / rate)
                elif not main_ws.cell(row=cell, column=35).value:
                    pricing.append(main_ws.cell(row=cell, column=35).value)
                elif "=" in main_ws.cell(row=cell, column=35).value:
                    logging.debug(f'Found1 = ')
                    pricing.append(f"{main_ws.cell(row=cell, column=35).value}/{rate}")  # Net Sales Price
                    logging.debug(f"{main_ws.cell(row=cell, column=35).value}/{rate}")
                else:
                    pricing.append(main_ws.cell(row=cell, column=35).value / rate)  # Net Sales Price
                logging.debug(f'NEXT1 ')

                if type(main_ws.cell(row=cell, column=53).value) in [type(1), type(0.1)]:  # To prevent the TypeError message at floats
                    logging.debug(f'NOTT1 Found = ')
                    pricing.append(main_ws.cell(row=cell, column=53).value / rate)  # Japan Margin Price
                elif not main_ws.cell(row=cell, column=53).value:
                    pricing.append(main_ws.cell(row=cell, column=53).value)
                elif "=" in main_ws.cell(row=cell, column=53).value:  # If a formula is in
                    logging.debug(f'Found2 = ')
                    pricing.append(f"{main_ws.cell(row=cell, column=53).value}/{rate}")  # Japan Margin Price
                    logging.debug(f"{main_ws.cell(row=cell, column=53).value}/{rate}")
                else:  # If not a float number and not a formula
                    logging.debug(f'NOTT1 Found = ')
                    pricing.append(main_ws.cell(row=cell, column=53).value / rate)  # Japan Margin Price
                logging.debug(f'Appended pricing data: {pricing} ')

                # Appending Volume Jan-Dec
                for vol_data in range(20, 32):
                    volume.append(main_ws.cell(row=cell, column=vol_data).value)
                logging.debug(f'Appended volume: {volume} ')

                # Appending Sales Jan- 1 month before current month
                for sales_data in range(38, 50):
                    if main_ws.cell(row=2, column=sales_data).value == month:
                        logging.debug(f"Appended until {month}. Empty cells must be filled with formula Price * Volume")
                        break

                    if not main_ws.cell(row=cell, column=sales_data).value:  # Cant divide a None value
                        sales.append(main_ws.cell(row=cell, column=sales_data).value)
                        logging.debug(f"Appended None {main_ws.cell(row=cell, column=sales_data).value}")
                    else:
                        sales.append(main_ws.cell(row=cell, column=sales_data).value / rate)
                        logging.debug(f"Appended Non None{main_ws.cell(row=cell, column=sales_data).value / rate}")
                logging.debug(f'Appended sales data: {sales} ')

                # Appending JM Jan - 1 month before current month
                for jm_data in range(56, 68):
                    if main_ws.cell(row=2, column=jm_data).value == month:
                        print(f"Appended until {month}. Empty cells must be filled with formula Margin * Volume")
                        break

                    if not main_ws.cell(row=cell, column=jm_data).value:
                        jm.append(main_ws.cell(row=cell, column=jm_data).value)
                    else:
                        jm.append(main_ws.cell(row=cell, column=jm_data).value / rate)
                logging.debug(f'Appended jm data: {jm} ')

            else:
                # Appending MTD data
                mtd_data.append(main_ws.cell(row=cell, column=19).value)  # MTD Volume
                mtd_data.append(main_ws.cell(row=cell, column=37).value)  # MTD Net Sales
                mtd_data.append(main_ws.cell(row=cell, column=55).value)  # MTD JM, usually 0 if not EOM data
                logging.debug(f'Appended mtd data: {mtd_data} ')

                # Appending NS Price and GM Price
                pricing.append(main_ws.cell(row=cell, column=35).value)  # Net Sales Price
                pricing.append(main_ws.cell(row=cell, column=53).value)  # Japan Margin Price
                logging.debug(f'Appended pricing data: {pricing} ')

                # Appending Volume Jan-Dec
                for vol_data in range(20, 32):
                    volume.append(main_ws.cell(row=cell, column=vol_data).value)
                logging.debug(f'Appended volume: {volume} ')

                # Appending Sales Jan- 1 month before current month
                for sales_data in range(38, 50):
                    if main_ws.cell(row=2, column=sales_data).value == month:
                        logging.debug(f"Appended until {month}. Empty cells must be filled with formula Price * Volume")
                        break

                    sales.append(main_ws.cell(row=cell, column=sales_data).value)
                logging.debug(f'Appended sales data: {sales} ')

                # Appending JM Jan - 1 month before current month
                for jm_data in range(56, 68):
                    if main_ws.cell(row=2, column=jm_data).value == month:
                        print(f"Appended until {month}. Empty cells must be filled with formula Margin * Volume")
                        break
                    jm.append(main_ws.cell(row=cell, column=jm_data).value)
                logging.debug(f'Appended jm data: {jm} ')

            # Writing the data to the new workbook
            offset = 3
            while True:
                logging.debug(f'Testing offset {offset} ')
                if not new_ws.cell(row=offset, column=1).value:

                    logging.debug(f'Empty cell found in row {offset} ')
                    # Writing main data
                    main_data_column = 1
                    for item in main_data:
                        new_ws.cell(row=offset, column=main_data_column).value = item
                        main_data_column += 1
                    logging.debug(f'Pasted main data')

                    # Writing MTD Data
                    new_ws.cell(row=offset, column=7).value = mtd_data[0]  # MTD Volume
                    new_ws.cell(row=offset, column=23).value = mtd_data[1]  # MTD Sales
                    new_ws.cell(row=offset, column=39).value = mtd_data[2]  # MTD JM
                    logging.debug(f'Pasted mtd data')

                    # Writing Pricing
                    new_ws.cell(row=offset, column=22).value = pricing[0]  # NS Price
                    new_ws.cell(row=offset, column=38).value = pricing[1]  # JM Price
                    logging.debug(f'Pasted pricing')

                    # Writing Volume
                    main_data_column = 8
                    for item in volume:
                        new_ws.cell(row=offset, column=main_data_column).value = item
                        main_data_column += 1
                    logging.debug(f'Pasted volume')

                    # Writing Sales
                    main_data_column = 24
                    for item in sales:
                        new_ws.cell(row=offset, column=main_data_column).value = item
                        main_data_column += 1

                    for fcell in range(main_data_column, 37):  # The cells after reporting month must contain formula
                        formula_column = get_column_letter(fcell - 16)
                        logging.warning(f"Formula column is {formula_column})")
                        new_ws.cell(row=offset, column=fcell).value = f"=SUM(V{offset}*{formula_column}{offset})"
                        logging.warning(f"Formula column is =SUM(AL{offset}*{formula_column}{offset})")
                    logging.debug(f'Pasted sales')

                    # Writing JM
                    main_data_column = 40
                    for item in jm:
                        new_ws.cell(row=offset, column=main_data_column).value = item
                        main_data_column += 1

                    for fcell in range(main_data_column, 52):  # The cells after reporting month must contain formula
                        formula_column = get_column_letter(fcell - 32)
                        logging.warning(f"Formula column is {formula_column})")
                        new_ws.cell(row=offset, column=fcell).value = f"=SUM(AL{offset}*{formula_column}{offset})"
                        logging.warning(f"Formula column is =SUM(AL{offset}*{formula_column}{offset})")
                    logging.debug(f'Pasted jm')

                    # Writing formulas for FY Forecast
                    new_ws.cell(row=offset, column=6).value = f"=SUM(H{offset}:S{offset})"  # FY Forecast Vol
                    new_ws.cell(row=offset, column=21).value = f"=SUM(X{offset}:AI{offset})"  # FY Forecast Sales
                    new_ws.cell(row=offset, column=37).value = f"=SUM(AN{offset}:AY{offset})"  # FY Forecast JM
                    logging.debug(f'Pasted FY Formulas')

                    # Writing formulas for YTD Actual Sales
                    for fcell in range(1, new_ws.max_column + 1):
                        logging.debug(f'Starting to define Columns for YTD')
                        if not new_ws.cell(row=1, column=fcell).value:  # Skips over None cells to prevent errors
                            continue

                        if "VOLUME" in new_ws.cell(row=1, column=fcell).value:
                            if new_ws.cell(row=2, column=fcell).value == month:
                                logging.debug(f'Found Volume column')
                                volume_column = get_column_letter(fcell - 1)  # YTD Actual excludes current month sales

                        elif "Sales" in new_ws.cell(row=1, column=fcell).value:
                            if new_ws.cell(row=2, column=fcell).value == month:
                                logging.debug(f'Found Sales column')
                                sales_column = get_column_letter(fcell - 1)

                        elif "JM" in new_ws.cell(row=1, column=fcell).value:
                            if new_ws.cell(row=2, column=fcell).value == month:
                                logging.debug(f'Found JM column')
                                jm_column = get_column_letter(fcell - 1)

                    logging.debug(f'Writing formulas for YTD Actual sales')
                    logging.warning(f"volume column is {volume_column}")
                    logging.warning(f"sales column is {sales_column}")
                    logging.warning(f"jm column is {jm_column}")

                    if month == "Jan":  # If Jan, there are no
                        new_ws.cell(row=offset, column=5).value = "XXX"
                        new_ws.cell(row=offset, column=20).value = "XXX"
                        new_ws.cell(row=offset, column=36).value = "XXX"
                    else:
                        new_ws.cell(row=offset, column=5).value = f"=SUM($H{offset}:{volume_column}{offset})"  # Volume
                        new_ws.cell(row=offset, column=20).value = f"=SUM($X{offset}:{sales_column}{offset})"  # Sales
                        new_ws.cell(row=offset, column=36).value = f"=SUM($AN{offset}:{jm_column}{offset})"  # JM
                    logging.debug(f'Pasted YTD Actual Sales')

                    break

                offset += 1
                logging.debug(f'Checking new row at row {offset}')

    # Hiding/grouping the passed months, so only the months ahead will be shown
    # We can use the column coordinates defined in the previous section
    # logging.warning(f"volume column is {volume_column}")
    # logging.warning(f"sales column is {sales_column}")
    # logging.warning(f"jm column is {jm_column}")
    #
    # new_ws.column_dimensions.group(start="H", end=volume_column, hidden=True)  # Hiding Volume Actuals
    # new_ws.column_dimensions.group(start="X", end=sales_column, hidden=True)  # Hiding Sales Actuals
    # new_ws.column_dimensions.group(start="AN", end=jm_column, hidden=True)  # Hiding JM Actuals

    # Summing the total amounts per month on the last row
    bold = Font(bold=True)
    fill = PatternFill(start_color='0099CC00', end_color='0099CC00', fill_type='solid')

    total_items = 2  # Row 1 and 2 are items too, but can be empty. To be error-safe, we just start at 3
    for item in range(3, new_ws.max_row):
        if new_ws.cell(row=item, column=1).value:
            total_items += 1
        else:
            break

        # We do not want the sums to be straight after the last item, we want 1 row in between
    new_ws.cell(row=total_items + 2, column=4).value = "TOTAL"
    new_ws.cell(row=total_items + 2, column=4).font = bold
    new_ws.cell(row=total_items+ 2, column=4).fill = fill

    for cell in range(5, 52):
        if new_ws.cell(row=2, column=cell).value == "Budget" or new_ws.cell(row=2, column=cell).value == "=V2":
            continue  # We do not want to sum the prices

        col = get_column_letter(cell)
        new_ws.cell(row=total_items + 2, column=cell).value = f"=SUM({col}3:{col}{total_items})"
        new_ws.cell(row=total_items + 2, column=cell).font = bold
        new_ws.cell(row=total_items + 2, column=cell).fill = fill
        new_ws.column_dimensions[col].width = 12  # Standardizing width so that it instantly is readable

    # Hiding/grouping the passed months, so only the months ahead will be shown
    # We can use the column coordinates defined in the YTD Actuals section
    logging.warning(f"volume column is {volume_column}")
    logging.warning(f"sales column is {sales_column}")
    logging.warning(f"jm column is {jm_column}")

    if month != "Jan":  # If January, it should not group to prevent errors
        new_ws.column_dimensions.group(start="H", end=volume_column, hidden=True)  # Hiding Volume Actuals
        new_ws.column_dimensions.group(start="X", end=sales_column, hidden=True)  # Hiding Sales Actuals
        new_ws.column_dimensions.group(start="AN", end=jm_column, hidden=True)  # Hiding JM Actuals

    # Writing YTD Column name to correct month, MTD to correct month
    logging.debug(f"Starting writing correct month for YTD and MTD")
    if month == "Jan":  # We cant go 1 month back from Jan, so we just put XXX
        YTD_date = "XXX"
        YTD_string = new_ws.cell(row=2, column=5).value
        new_ws.cell(row=2, column=5).value = re.sub("\d\d\d\d\w\w\w$", f"{datetime.datetime.now().year}{YTD_date}",
                                                    YTD_string)
    else:
        YTD_date = datetime.datetime.strptime(month, "%b")  # Going 1 month back as YTD does not count current month
        YTD_date = YTD_date.replace(month=YTD_date.month - 1)
        YTD_date = YTD_date.strftime("%b")

        YTD_string = new_ws.cell(row=2, column=5).value
        new_ws.cell(row=2, column=5).value = re.sub("\d\d\d\d\w\w\w$", f"{datetime.datetime.now().year}{YTD_date}", YTD_string)

    new_ws.cell(row=2, column=7).value = re.sub("\w\w\w$", month, new_ws.cell(row=2, column=7).value)


    logging.debug(f'No errors, returning workbook')
    return new_wb


























