# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'ltoSQLTableEnlarged.ui'
#
# Created by: PyQt5 UI code generator 5.13.2
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets, Qt
from PyQt5.QtSql import QSqlDatabase, QSqlTableModel, QSqlQuery
import os
import sqlite3
from openpyxl.utils import get_column_letter
from UpdateSalesDatabase import *
from SalesResultsDashboard import *



db = QSqlDatabase("QSQLITE")
db.setDatabaseName("Databases\\sales.db")
db.open()

class Ui_sales_database(object):
    def setupUi(self, sales_database):
        sales_database.setObjectName("sales_database")
        sales_database.resize(1600, 805)
        sales_database.setMinimumSize(QtCore.QSize(1600, 800))
        sales_database.setMaximumSize(QtCore.QSize(16777215, 16777215))
        sales_database.setStyleSheet("background-color: rgb(0, 40, 72);")
        self.centralwidget = QtWidgets.QWidget(sales_database)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.setObjectName("verticalLayout")
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setMinimumSize(QtCore.QSize(1601, 80))
        self.frame.setStyleSheet("background-color: rgb(0, 35, 72);")
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.logo_label = QtWidgets.QLabel(self.frame)
        self.logo_label.setGeometry(QtCore.QRect(0, 0, 101, 81))
        self.logo_label.setText("")
        self.logo_label.setPixmap(QtGui.QPixmap("Images/DoehlerLogo.png"))
        self.logo_label.setScaledContents(True)
        self.logo_label.setObjectName("logo_label")
        self.app_title = QtWidgets.QLabel(self.frame)
        self.app_title.setGeometry(QtCore.QRect(100, 0, 1491, 81))
        self.app_title.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("MS Reference Sans Serif")
        font.setPixelSize(32)
        font.setBold(False)
        font.setWeight(50)
        font.setKerning(True)
        self.app_title.setFont(font)
        self.app_title.setStyleSheet("background-color: rgb(0, 35, 72);\n"
                                     "color: rgb(255, 255, 255);")
        self.app_title.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.app_title.setFrameShadow(QtWidgets.QFrame.Plain)
        self.app_title.setMidLineWidth(0)
        self.app_title.setTextFormat(QtCore.Qt.PlainText)
        self.app_title.setAlignment(QtCore.Qt.AlignCenter)
        self.app_title.setWordWrap(False)
        self.app_title.setObjectName("app_title")
        self.verticalLayout.addWidget(self.frame)
        self.actions_frame = QtWidgets.QFrame(self.centralwidget)
        self.actions_frame.setMinimumSize(QtCore.QSize(1601, 71))
        self.actions_frame.setStyleSheet("background-color: rgb(0, 40, 72);")
        self.actions_frame.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.actions_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.actions_frame.setObjectName("actions_frame")
        self.query_lineedit = QtWidgets.QLineEdit(self.actions_frame)
        self.query_lineedit.setGeometry(QtCore.QRect(10, 30, 711, 31))
        font = QtGui.QFont()
        font.setPixelSize(15)
        self.query_lineedit.setFont(font)
        self.query_lineedit.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.query_lineedit.setObjectName("query_lineedit")
        self.query_label = QtWidgets.QLabel(self.actions_frame)
        self.query_label.setGeometry(QtCore.QRect(10, 0, 261, 21))
        font = QtGui.QFont()
        font.setPixelSize(20)
        font.setBold(True)
        font.setUnderline(False)
        font.setWeight(75)
        font.setKerning(True)
        self.query_label.setFont(font)
        self.query_label.setStyleSheet("color: rgb(255, 255, 255);")
        self.query_label.setObjectName("query_label")

        self.visualize_button = QtWidgets.QPushButton(self.actions_frame)
        self.visualize_button.setGeometry(QtCore.QRect(850, 10, 181, 51))
        font = QtGui.QFont()
        font.setPixelSize(15)
        self.visualize_button.setFont(font)
        self.visualize_button.setStyleSheet("color: rgb(255, 255, 255);\n"
                                            "background-color: rgb(0, 111, 196);")
        self.visualize_button.setObjectName("visualize_button")

        self.extract_button = QtWidgets.QPushButton(self.actions_frame)
        self.extract_button.setGeometry(QtCore.QRect(740, 10, 100, 51))
        font = QtGui.QFont()
        font.setPixelSize(10)
        self.extract_button.setFont(font)
        self.extract_button.setStyleSheet("color: rgb(255, 255, 255);\n"
                                            "background-color: rgb(0, 111, 196);")
        self.extract_button.setObjectName("visualize_button")

        self.update_button = QtWidgets.QPushButton(self.actions_frame)
        self.update_button.setGeometry(QtCore.QRect(1040, 10, 181, 51))
        font = QtGui.QFont()
        font.setPixelSize(15)
        self.update_button.setFont(font)
        self.update_button.setStyleSheet("color: rgb(255, 255, 255);\n"
                                         "background-color: rgb(0, 111, 196);")
        self.update_button.setObjectName("update_button")
        self.update_label = QtWidgets.QLabel(self.actions_frame)
        self.update_label.setEnabled(True)
        self.update_label.setGeometry(QtCore.QRect(1230, 20, 361, 31))
        font = QtGui.QFont()
        font.setPixelSize(10)
        self.update_label.setFont(font)
        self.update_label.setStyleSheet("color: rgb(255, 255, 255);\n"
                                        "border-color: rgb(255, 255, 255);\n"
                                        "background-color: rgb(0, 85, 127);")
        self.update_label.setObjectName("update_label")
        self.tableload_label = QtWidgets.QLabel(self.actions_frame)
        self.tableload_label.setGeometry(QtCore.QRect(570, 0, 171, 21))
        font = QtGui.QFont()
        font.setPixelSize(20)
        font.setBold(True)
        font.setUnderline(False)
        font.setWeight(75)
        font.setKerning(True)
        self.tableload_label.setFont(font)
        self.tableload_label.setStyleSheet("color: rgb(255, 255, 255);")
        self.tableload_label.setObjectName("tableload_label")
        self.verticalLayout.addWidget(self.actions_frame)
        self.table_data = QtWidgets.QTableView()
        self.table_data.setMinimumSize(QtCore.QSize(1581, 611))
        self.table_data.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.table_data.setObjectName("table_data")
        font = QtGui.QFont()
        font.setPixelSize(14)
        self.table_data.setFont(font)
        self.verticalLayout.addWidget(self.table_data)
        sales_database.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(sales_database)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1600, 18))
        self.menubar.setObjectName("menubar")
        sales_database.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(sales_database)
        self.statusbar.setObjectName("statusbar")
        sales_database.setStatusBar(self.statusbar)

        self.retranslateUi(sales_database)
        QtCore.QMetaObject.connectSlotsByName(sales_database)

        self.model = QSqlTableModel(db=db)
        self.model.setTable("sales")
        self.model.setEditStrategy(QSqlTableModel.OnManualSubmit)

        self.query = QSqlQuery(db=db)
        self.query.prepare("SELECT * FROM sales ORDER BY month, code, customer, bu1, bu2") # setSort not working, have to do through Query
        self.query.exec_()
        self.model.setQuery(self.query)

        self.table_data.setModel(self.model)
        self.table_data.resizeColumnsToContents()

        self.update_button.clicked.connect(self.update_lto_clicked)
        self.query_lineedit.returnPressed.connect(self.update_query)
        self.extract_button.clicked.connect(self.extract_query)
        self.visualize_button.clicked.connect(self.visualize_clicked)

    def retranslateUi(self, sales_database):
        _translate = QtCore.QCoreApplication.translate
        sales_database.setWindowTitle(_translate("sales_database", "MainWindow"))
        self.app_title.setText(_translate("sales_database", "Sales Results EURk - Doehler Japan"))
        self.query_label.setText(_translate("sales_database", "Query Box (SQLite syntax)"))
        self.visualize_button.setText(_translate("sales_database", "Visualize displayed data"))
        self.extract_button.setText(_translate("sales_database,", "Extract shown data"))
        self.update_button.setText(_translate("sales_database", "Update database"))
        self.update_label.setText(
            _translate("sales_database", "To update  - Please select excel file originating from \'WE LEAD ANALYTICS\'"))
        self.tableload_label.setText(_translate("sales_database", "sales table loaded"))

    def update_lto_clicked(self):
        self.SalesResultsUploadWindow = QtWidgets.QMainWindow()
        self.ui = Ui_SalesResultsUploadWindow()
        self.ui.setupUi(self.SalesResultsUploadWindow)
        self.SalesResultsUploadWindow.show()

    def visualize_clicked(self):
        self.month_visualize = QtWidgets.QMainWindow()
        self.ui = Ui_VisualizeMonth()
        self.ui.setupUi(self.month_visualize)
        self.month_visualize.show()

    def update_query(self):
        if not self.query_lineedit.text():  # No input basically means the user wants to refresh the database
            db.close()  # In order to refresh, we close and re-open the database
            db.open()
            self.model = QSqlTableModel(db=db)
            self.model.setTable("sales")
            # self.model.setEditStrategy(QSqlTableModel.OnRowChange)

            self.query.prepare("SELECT * FROM sales ORDER BY month, code, customer, bu1, bu2")
            self.query.exec_()
            self.model.setQuery(self.query)

            self.table_data.setModel(self.model)
            self.table_data.resizeColumnsToContents()
        else:
            self.query.prepare(self.query_lineedit.text())

            self.query.exec_()
            self.model.setQuery(self.query)
            self.table_data.resizeColumnsToContents()

    def extract_query(self):
        self.extract_dir = QtWidgets.QFileDialog.getSaveFileName(filter="*.xlsx")

        if not self.extract_dir[0]:
            return

        connect = sqlite3.connect("Databases\\sales.db")
        c = connect.cursor()
        print("Connected")

        if not self.query_lineedit.text():
            print("No query, setting default")
            extract_query = "SELECT * FROM sales ORDER BY month, code, customer, bu1, bu2"
        else:
            print("Query found")
            extract_query = self.query_lineedit.text()

        c.execute(extract_query)
        extract_data = c.fetchall()
        column_headers = c.description  # c.description displays the column header in tuple ('column', None, None)
        # Which is why it is important to only grab the [0][0] from the column_Headers.

        self.database_columns = [column_header[0] for column_header in column_headers]  # Grabbing [0][0] and appending
        print(self.database_columns)

        extract = pyxl.Workbook()
        extract_ws = extract.active

        # Creating headers
        for index, data in enumerate(self.database_columns):
            extract_ws.cell(row=1, column=index + 1).value = data
            print(data)

        column_width = []
        for index, data in enumerate(extract_data):
            for index2, data2 in enumerate(data):
                extract_ws.cell(row=index + 2, column=index2 + 1).value = data2  # +2 is because we start in row 2
                if len(column_width) > index2:  # If the length of list is smaller than index, then we stil lhave to add column data
                    if len(str(data2)) > column_width[index2]:  #
                        column_width[index2] = len(str(data2))
                else:
                    column_width.append(len(str(data2)))

        for index, column_sizing in enumerate(column_width):
            extract_ws.column_dimensions[get_column_letter(index + 1)].width = column_sizing * 1.2

        print(column_width)

        extract.save(self.extract_dir[0])

        # with open(self.extract_dir[0], "w", newline="", errors="ignore") as new:
        #     print("opened")
        #     writer = csv.writer(new, delimiter=";")
        #     writer.writerow(self.database_columns)
        #     for line in extract_data:
        #         writer.writerow(line)



class Ui_SalesResultsUploadWindow(object):
    def setupUi(self, SalesResultsUploadWindow):
        SalesResultsUploadWindow.setObjectName("SalesResultsUploadWindow")
        SalesResultsUploadWindow.resize(1379, 620)
        SalesResultsUploadWindow.setMinimumSize(QtCore.QSize(1379, 620))
        SalesResultsUploadWindow.setMaximumSize(QtCore.QSize(1379, 620))
        self.centralwidget = QtWidgets.QWidget(SalesResultsUploadWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.example_picture = QtWidgets.QLabel(self.centralwidget)
        self.example_picture.setGeometry(QtCore.QRect(0, 360, 1381, 201))
        self.example_picture.setText("")
        self.example_picture.setPixmap(QtGui.QPixmap("Images/sales_results_new_example.png"))
        self.example_picture.setObjectName("example_picture")
        self.upload_box = QtWidgets.QGroupBox(self.centralwidget)
        self.upload_box.setGeometry(QtCore.QRect(9, 79, 621, 261))
        font = QtGui.QFont()
        font.setPixelSize(30)
        self.upload_box.setFont(font)
        self.upload_box.setObjectName("upload_box")
        self.month_combobox = QtWidgets.QComboBox(self.upload_box)
        self.month_combobox.setGeometry(QtCore.QRect(20, 50, 171, 31))
        font = QtGui.QFont()
        font.setPixelSize(18)
        self.month_combobox.setFont(font)
        self.month_combobox.setObjectName("month_combobox")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_combobox.addItem("")
        self.month_label = QtWidgets.QLabel(self.upload_box)
        self.month_label.setGeometry(QtCore.QRect(230, 50, 291, 31))
        font = QtGui.QFont()
        font.setPixelSize(20)
        self.month_label.setFont(font)
        self.month_label.setObjectName("month_label")
        self.upload1_lbutton = QtWidgets.QPushButton(self.upload_box)
        self.upload1_lbutton.setGeometry(QtCore.QRect(20, 100, 171, 61))
        font = QtGui.QFont()
        font.setPixelSize(18)
        self.upload1_lbutton.setFont(font)
        self.upload1_lbutton.setObjectName("upload1_lbutton")
        self.upload1_label = QtWidgets.QLabel(self.upload_box)
        self.upload1_label.setGeometry(QtCore.QRect(230, 110, 351, 31))
        font = QtGui.QFont()
        font.setPixelSize(20)
        self.upload1_label.setFont(font)
        self.upload1_label.setFrameShape(QtWidgets.QFrame.Panel)
        self.upload1_label.setObjectName("upload1_label")
        self.upload2_button = QtWidgets.QPushButton(self.upload_box)
        self.upload2_button.setGeometry(QtCore.QRect(20, 190, 171, 61))
        font = QtGui.QFont()
        font.setPixelSize(18)
        self.upload2_button.setFont(font)
        self.upload2_button.setObjectName("upload2_button")
        self.upload2_label = QtWidgets.QLabel(self.upload_box)
        self.upload2_label.setGeometry(QtCore.QRect(230, 200, 351, 31))
        font = QtGui.QFont()
        font.setPixelSize(20)
        self.upload2_label.setFont(font)
        self.upload2_label.setFrameShape(QtWidgets.QFrame.Panel)
        self.upload2_label.setObjectName("upload2_label")
        self.explanation_box = QtWidgets.QGroupBox(self.centralwidget)
        self.explanation_box.setGeometry(QtCore.QRect(650, 80, 721, 261))
        font = QtGui.QFont()
        font.setPixelSize(30)
        self.explanation_box.setFont(font)
        self.explanation_box.setObjectName("explanation_box")
        self.explanation_browser = QtWidgets.QTextBrowser(self.explanation_box)
        self.explanation_browser.setGeometry(QtCore.QRect(30, 40, 671, 211))
        self.explanation_browser.setObjectName("explanation_browser")
        self.app_title = QtWidgets.QLabel(self.centralwidget)
        self.app_title.setGeometry(QtCore.QRect(100, 0, 1281, 81))
        self.app_title.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("MS Reference Sans Serif")
        font.setPixelSize(32)
        font.setBold(False)
        font.setWeight(50)
        font.setKerning(True)
        self.app_title.setFont(font)
        self.app_title.setStyleSheet("background-color: rgb(0, 35, 72);\n"
"color: rgb(255, 255, 255);")
        self.app_title.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.app_title.setFrameShadow(QtWidgets.QFrame.Plain)
        self.app_title.setMidLineWidth(0)
        self.app_title.setTextFormat(QtCore.Qt.PlainText)
        self.app_title.setAlignment(QtCore.Qt.AlignCenter)
        self.app_title.setWordWrap(False)
        self.app_title.setObjectName("app_title")
        self.doehler_label = QtWidgets.QLabel(self.centralwidget)
        self.doehler_label.setGeometry(QtCore.QRect(0, 0, 101, 81))
        self.doehler_label.setText("")
        self.doehler_label.setPixmap(QtGui.QPixmap("Images/DoehlerLogo.png"))
        self.doehler_label.setScaledContents(True)
        self.doehler_label.setObjectName("doehler_label")
        SalesResultsUploadWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(SalesResultsUploadWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1379, 18))
        self.menubar.setObjectName("menubar")
        SalesResultsUploadWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(SalesResultsUploadWindow)
        self.statusbar.setObjectName("statusbar")
        SalesResultsUploadWindow.setStatusBar(self.statusbar)

        self.retranslateUi(SalesResultsUploadWindow)
        QtCore.QMetaObject.connectSlotsByName(SalesResultsUploadWindow)

        self.upload1_lbutton.clicked.connect(self.upload_sales)

    def retranslateUi(self, SalesResultsUploadWindow):
        _translate = QtCore.QCoreApplication.translate
        SalesResultsUploadWindow.setWindowTitle(_translate("SalesResultsUploadWindow", "MainWindow"))
        self.upload_box.setTitle(_translate("SalesResultsUploadWindow", "Upload SAP file"))
        self.month_combobox.setItemText(0, _translate("SalesResultsUploadWindow", "Select"))
        self.month_combobox.setItemText(1, _translate("SalesResultsUploadWindow", "01"))
        self.month_combobox.setItemText(2, _translate("SalesResultsUploadWindow", "02"))
        self.month_combobox.setItemText(3, _translate("SalesResultsUploadWindow", "03"))
        self.month_combobox.setItemText(4, _translate("SalesResultsUploadWindow", "04"))
        self.month_combobox.setItemText(5, _translate("SalesResultsUploadWindow", "05"))
        self.month_combobox.setItemText(6, _translate("SalesResultsUploadWindow", "06"))
        self.month_combobox.setItemText(7, _translate("SalesResultsUploadWindow", "07"))
        self.month_combobox.setItemText(8, _translate("SalesResultsUploadWindow", "08"))
        self.month_combobox.setItemText(9, _translate("SalesResultsUploadWindow", "09"))
        self.month_combobox.setItemText(10, _translate("SalesResultsUploadWindow", "10"))
        self.month_combobox.setItemText(11, _translate("SalesResultsUploadWindow", "11"))
        self.month_combobox.setItemText(12, _translate("SalesResultsUploadWindow", "12"))
        self.month_label.setText(_translate("SalesResultsUploadWindow", "Month of sales result data"))
        self.upload1_lbutton.setText(_translate("SalesResultsUploadWindow", "Upload Sales Result"))
        self.upload1_label.setText(_translate("SalesResultsUploadWindow", "No file loaded..."))
        self.upload2_button.setText(_translate("SalesResultsUploadWindow", "Upload Budget"))
        self.upload2_label.setText(_translate("SalesResultsUploadWindow", "Optional..."))
        self.explanation_box.setTitle(_translate("SalesResultsUploadWindow", "Upload explanation"))
        self.explanation_browser.setHtml(_translate("SalesResultsUploadWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS UI Gothic\'; font-size:30pt; font-weight:400; font-style:normal;\">\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600;\">Please use the variables shown in the image below:</span></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600;\"><br /></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600;\">Filter</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- Company code</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- Rep. Customer</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- B2B Level 1</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- B2B Level 2</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- Material (1-9)</span></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600;\"><br /></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600;\">Key Figures</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- Sales Volume</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- Gross Sales</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- Net Sales Ex Works</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- Standard GM</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- Standard CM1</span></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600;\"><br /></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600;\">Period Structure</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- *1.000</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt;\">- MM.YYYY &gt; 1 Month only</span></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS UI Gothic\'; font-size:18pt;\"><br /></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600; text-decoration: underline;\">Have the columns ordered as shown in the image below</span></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600; text-decoration: underline;\"><br /></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600; text-decoration: underline;\">Be sure to have no cells merged. Unmerge all cells.</span></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS UI Gothic\'; font-size:18pt; font-weight:600;\"><br /></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS UI Gothic\'; font-size:18pt;\"><br /></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS UI Gothic\';\"><br /></p></body></html>"))
        self.app_title.setText(_translate("SalesResultsUploadWindow", "Doehler Japan - Upload Sales Results"))


    def upload_sales(self):

        # Making sure a month is selected
        self.month = self.month_combobox.currentText()
        if self.month == "Select":
            error_dialog = QtWidgets.QErrorMessage()
            error_dialog.showMessage("No month has been selected. Please select a month")
            error_dialog.show()
            error_dialog.exec_()
            return

        # Checking whether this month already is in the database. If it is, give the user the option to overwrite
        month_in = self.month_check()

        if month_in != 0:
            message_dialog = QtWidgets.QMessageBox()
            message_dialog.setText(f"{month_in} lines of data for month {self.month} are already in the database.\n"
                                   f"Do you want to overwrite these lines with new sales results data?")
            message_dialog.setStandardButtons(QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel)
            message_dialog_pressed = message_dialog.exec_()

            if message_dialog_pressed != QtWidgets.QMessageBox.Ok:
                message_dialog.setText(f"Upload canceled.")
                message_dialog.setStandardButtons(QtWidgets.QMessageBox.Ok)
                message_dialog.exec_()
                return
            else:
                self.overwrite_month()  # If Ok was pressed, we overwrite.

        # Obtaining the file path from the file dialog
        self.sales_dir = QtWidgets.QFileDialog.getOpenFileName(filter="*xlsx")
        if not self.sales_dir[0]:
            self.upload1_label.setText("File was not selected.")
            self.upload1_label.setStyleSheet("background-color: rgb(255, 0, 0);")
            return

        self.sales_dir = os.path.abspath(self.sales_dir[0])
        self.upload1_label.setText(f"{os.path.basename(self.sales_dir)} is currently loaded")
        self.upload1_label.setStyleSheet("background-color: rgb(174, 217, 167);")

        # Testing the template of the file, whether it is eligible for upload to db
        # Todo - Connect the functions of UpdateSalesDatabase.py to here.
        template_test = UpdateSalesResultsTest(self.sales_dir, self.month)

        # template_test returns one of: "Header failed", "Header 2 failed", "Header 3 failed", dictionary of headers
        if isinstance(template_test, str):  # Test has failed if this gets read
            error_dialog = QtWidgets.QErrorMessage()
            error_dialog.showMessage(f"There seems to be a mismatch in the template:\n"
                                     f"{template_test}")
            error_dialog.exec_()
            return

        #Test is OK, we received a dictionary value with the header co-ordinates.
        upload_to_db = create_sales_db(self.sales_dir, self.month, template_test)
        print("Exited the poop")
        print(upload_to_db)

        message_dialog = QtWidgets.QMessageBox()
        message_dialog.setText(f"Writing completed. {upload_to_db} new lines have been added to the database.")
        message_dialog.setStandardButtons(QtWidgets.QMessageBox.Ok)
        message_dialog.exec_()


    def month_check(self):
        connect = sqlite3.connect("Databases\\sales.db")
        c = connect.cursor()
        print('Connected to db')

        c.execute(f"SELECT COUNT(month) FROM sales WHERE month = '{self.month}'")
        print("Succesfully executed")

        data_amt = c.fetchall()
        print(data_amt[0][0])

        c.close()
        connect.close()

        return data_amt[0][0]

    def overwrite_month(self):

        connect = sqlite3.connect("Databases\\sales.db")
        c = connect.cursor()
        c.execute(f"DELETE FROM sales WHERE month = '{self.month}'")

        connect.commit()
        c.close()
        connect.close()


class Ui_VisualizeMonth(object):
    def setupUi(self, VizualizeMonth):
        VizualizeMonth.resize(200, 200)
        VizualizeMonth.setMinimumSize(QtCore.QSize(250, 100))
        VizualizeMonth.setMaximumSize(QtCore.QSize(250, 100))

        self.layout = QtWidgets.QVBoxLayout()

        self.combo = QtWidgets.QComboBox()
        self.combo.addItems(["Select Month for visuals", "01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"])
        self.combo.setStyleSheet("background-color: rgb(174, 217, 167);")

        self.combo2 = QtWidgets.QComboBox()
        self.combo2.addItems(["Select 'Local' or 'Destination' results", "Local Sales Results", "Destination Sales Results"])
        self.combo2.setStyleSheet("background-color: rgb(100, 217, 167);")

        self.button = QtWidgets.QPushButton("Proceed with visualization")

        self.layout.addWidget(self.combo)
        self.layout.addWidget(self.combo2)
        self.layout.addWidget(self.button)

        self.widget = QtWidgets.QWidget()
        self.widget.setLayout(self.layout)
        VizualizeMonth.setCentralWidget(self.widget)

        self.button.clicked.connect(self.show_data)

    def show_data(self):
        if self.combo.currentText() == "Select Month for visuals":
            message = QtWidgets.QMessageBox()
            message.setText("You did not select a month.\n\nPlease select a month.")
            message.setStandardButtons(QtWidgets.QMessageBox.Ok)
            message.exec_()
            return

        if self.combo2.currentText() == "Select 'Local' or 'Destination' results":
            message = QtWidgets.QMessageBox()
            message.setText("Please select what sales results should be visualized.\n\nPlease select 'Local' or Destination.")
            message.setStandardButtons(QtWidgets.QMessageBox.Ok)
            message.exec_()
            return

        # Checking whether the month is in the database
        month_in_database = self.check_month()
        if not month_in_database:
            message = QtWidgets.QMessageBox()
            message.setText("This month has not been uploaded to the database yet.\n\n"
                            "Please upload the month or select a different month.")
            message.setStandardButtons(QtWidgets.QMessageBox.Ok)
            message.exec_()
            return

        self.show_dashboard()


    def show_dashboard(self):
        self.SalesResultsVis = QtWidgets.QMainWindow()
        self.ui = Ui_SalesResultsVis()
        self.ui.setupUi(self.SalesResultsVis, self.combo.currentText(), self.combo2.currentText())
        self.SalesResultsVis.show()




    def check_month(self):
        month = self.combo.currentText()

        connect = sqlite3.connect("Databases\\sales.db")
        c = connect.cursor()

        c.execute(f"SELECT COUNT(month) FROM sales WHERE month = '{month}'")
        data_amt = c.fetchall()

        c.close()
        connect.close()

        if data_amt[0][0] == 0:
            return False

        return True






if __name__ == "__main__":
    import sys
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")
    sales_database = QtWidgets.QMainWindow()
    ui = Ui_sales_database()
    ui.setupUi(sales_database)
    sales_database.show()
    sys.exit(app.exec_())
