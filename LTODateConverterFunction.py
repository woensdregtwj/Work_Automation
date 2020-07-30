import openpyxl as pyxl
from openpyxl.styles import *
import datetime
import re


def format_test1(sheet):
    test_sheet = sheet

    # For if the file had already been formatted, meaning we have to check the headers in row 1
    test_column_headers_formatted = ["Opportunity ID", "Opportunity Name", "Customer", "Region", "Area",
                                     "Customer Responsibility", "Döhler Opportunity", "Customer Opportunity",
                                     "Opportunity Stage", "Stage Start Date", "Start Date", "Launch Date",
                                     "Diamond Project", "Project Importance", "Development Order ID",
                                     "Development Order Status", "Potential (EURk)", "Volume in k"]

    # If file is not formatted, then row 1 should be empty, and the headers should be in row 2
    test_column_headers_unformatted = ["Opportunity ID", "Opportunity Name", "Customer", "Region", "Area",
                                       "Customer Responsibility", "Döhler Opportunity", "Customer Opportunity",
                                       "Opportunity Stage", "Stage Start Date", "Start Date", "Launch Date",
                                       "Diamond Project", "Project Importance", "Development Order ID",
                                       "Development Order Status", "EUR", ""]

    # Test 1, checking headers
    if test_sheet["A1"].value:
        for index, header in enumerate(test_column_headers_formatted):
            if header == test_sheet.cell(row=1, column=index + 1).value:
                print("Header correct, data for this column should be OK")
            else:
                return False
    else:
        for index, header in enumerate(test_column_headers_unformatted):
            if header == test_sheet.cell(row=2, column=index + 1).value:
                if test_sheet.cell(row=3, column=1).value == "Totals":
                    print("Header correct, data for this column should be OK")
            else:
                return False

    return True

def lto_date_format(lto_file):
    lto_wb = pyxl.load_workbook(lto_file)
    lto_ws = lto_wb.active

    is_correct_template = format_test1(lto_ws)
    if not is_correct_template:
        print("FAILED TEST")
        return False  # Test failed, template is not correct, returning this to GUI
    print("Passed the test, proceeding with formatting!")

    if not lto_ws["A1"].value:
        lto_ws.delete_rows(1)
        lto_ws.delete_rows(2)
        lto_ws["Q1"].value = "Potential (EURk)"
        lto_ws["R1"].value = "Volume in k"
        lto_ws.auto_filter.ref = lto_ws.dimensions

    for cell in range(2, lto_ws.max_row + 1):
        for column in range(10, 13):
            try:
                date = datetime.datetime.strptime(lto_ws.cell(row=cell, column=column).value, "%d.%m.%Y")
                lto_ws.cell(row=cell, column=column).value = date.strftime("%Y/%m/%d")
            except:
                if re.search("\d\d\d\d", lto_ws.cell(row=cell, column=column).value):
                    print(f"Date already was correct - {lto_ws.cell(row=cell, column=column).value} \
                     {lto_ws.cell(row=cell, column=1).value}")
                else:
                    print(f"Expected a date, but found {lto_ws.cell(row=cell, column=column).value}. \
                     Marking the cell red for checking.")
                    redFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
                    lto_ws.cell(row=cell, column=column).fill = redFill
                continue

    for col in ["D", "E", "G", "H", "J", "K", "M", "N", "O", "P"]:
        if not lto_ws.column_dimensions[col].hidden:
            lto_ws.column_dimensions[col].hidden = True

    return lto_wb
