#! python3

# At EOM, we always have to manually adjust the layout so that it reads the correct cell for the months.


import pyautogui
import re
import os
import pprint
import copy
import sys
import openpyxl as pyxl
import datetime
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Data structures:
sales_data = {}  # Our data structure
storage_delivery_trade = {'delivery_storage': 0,
                          'trade_volume': 0,
                          'trade_sales': 0,
                          'trade_profit': 0}  # for storing the total net sales of all storage data and trade business

trade_customers = ['Moriyama']  # Add here additional trade customers if they will ever come

# Section 1 - Getting the path and file based on user input - Sales Result file and Sales FC file

# 1.1 Getting Sales Result file
while True:
    reporting_month = pyautogui.prompt("Enter month for forecast as format MMM. E.g. 'Jul'")
    if not reporting_month:
        pyautogui.alert('Quitting program.')
        sys.exit()
    try:
        datetime.datetime.strptime(reporting_month, '%b')  # Testing whether the month format is correct
    except ValueError:
        pyautogui.alert('You did not correctly enter the month as format MMM. Please try again.')
    break

while True:
    filepath = pyautogui.prompt("Insert file directory for ===Sales Results===\n")

    if not filepath:
        pyautogui.alert('Quitting program.')
        sys.exit()

    sales_folder = os.path.abspath(filepath)
    if not Path(sales_folder).exists():
        pyautogui.alert('Given path does not exist, please try again.')
        continue
    break

for file in Path(sales_folder).glob('*.xlsx'):
    if re.search(r'((S|s)ales) ((R|r)esult|(U|u)pload)', os.path.basename(file)) and '~' not in os.path.basename(file):
        confirm = pyautogui.confirm(f"Is '{os.path.basename(file)}' the correct file?", buttons=['Yes', 'No', 'QUIT'])
        if confirm == 'Yes':
            sales_file = os.path.abspath(file)
            break
        elif confirm == 'QUIT':
            pyautogui.alert('Quitting program')
            sys.exit()
        sales_file = None  # Used for next if statement to mention that no sales file has been selected/found
        continue
try:
    if not sales_file:
        pyautogui.alert('You have not selected any sales upload file. Please check path.')
        sys.exit()
except NameError:
    pyautogui.alert('No Sales Result or Sales Upload file was found in this file path.')
    sys.exit()


# 1.2 Getting Sales FC file
while True:
    filepath = pyautogui.prompt("Insert file directory for ===Forecast===\n")

    if not filepath:
        pyautogui.alert('Quitting program.')
        sys.exit()

    forecast_folder = os.path.abspath(filepath)
    if not Path(sales_folder).exists():
        pyautogui.alert('Given path does not exist, please try again.')
        continue
    break

# filepath = pyautogui.prompt("Insert file directory for ===Forecast===")

# forecast_folder = os.path.abspath(filepath)

for file in Path(forecast_folder).glob('*.xlsx'):
    if re.search(f'Sales FC {reporting_month}.*', str(file)) and '~' not in str(file):
        confirm = pyautogui.confirm(f"Is '{os.path.basename(file)}' the correct file?", buttons=['Yes', 'No', 'QUIT'])
        if confirm == 'Yes':
            forecast_file = os.path.abspath(file)
            break
        elif confirm == 'QUIT':
            pyautogui.alert('Quitting program')
            sys.exit()
        forecast_file = None
        continue

try:
    if not forecast_file:
        pyautogui.alert('You have not selected any Forecast File. Please check path.')
        sys.exit()
except NameError:
    pyautogui.alert("No forecast file found in this path. Check for a file name similar to:\n"
                    "'Sales FC Mon' >>> Sales FC Jul")

# Section 2 - Reading into the files and creating data structures

sales_wb = pyxl.load_workbook(sales_file)
for sheet in sales_wb.sheetnames:
    if not sales_wb[sheet]['A1'].value:
        continue
    elif "売上" in sales_wb[sheet]['A1'].value:
        sales_sheet = sales_wb[sheet]
        break

for cell in range(4, sales_sheet.max_row + 1):
    customer = sales_sheet.cell(row=cell, column=1).value
    productID = sales_sheet.cell(row=cell, column=7).value
    product = sales_sheet.cell(row=cell, column=9).value
    volume = sales_sheet.cell(row=cell, column=10).value
    sales = sales_sheet.cell(row=cell, column=12).value

    # Now start filling up the dictionary with /getdefault

    sales_data.setdefault(customer, {})
    sales_data[customer].setdefault(productID, {})
    sales_data[customer][productID].setdefault(product, {'volume': 0, 'sales': 0, 'profit': 0})

    sales_data[customer][productID][product]['volume'] += round(float(volume), 2)
    sales_data[customer][productID][product]['sales'] += int(sales)

    if not sales_sheet.cell(row=cell, column=13).value:  # if no cogs in file, then we just put 0 as profit
        sales_data[customer][productID][product]['profit'] += 0
    elif type(sales_sheet.cell(row=cell, column=13).value) == type(str()):  # this has 'NO COGS' message, profit = sales
        sales_data[customer][productID][product]['profit'] += int(sales)
    else:  # has cogs, we calculate the profit by sales - cogs
        sales_data[customer][productID][product]['profit'] += int(sales) - sales_sheet.cell(row=cell, column=13).value

# Section 3. Pasting data from the data structure into the designated cells of the correct month on the forecast sheet

forecast_wb = pyxl.load_workbook(forecast_file)
forecast_sheet = forecast_wb['2020 All']

unmatched_sales = copy.deepcopy(sales_data)  # we want to keep the original sales_data dict as it is.
# for this dict, we will take out all succesfull links and only keep the unmatched sales so that we can check manually
# and input it into the sales forecast ourself

# In comparison to the 'Forecast Updater.py', this time we have to find the correct column.
# We can find it by checking whether e.g. row 1 value == 'SALES VOLUME' and row 2 value == forecast_month

for cell in range(1, forecast_sheet.max_column):
     if re.search(r'SALES.*VOLUME', str(forecast_sheet.cell(row=1, column=cell).value)):
         if forecast_sheet.cell(row=2, column=cell).value == reporting_month:
             print(f'{cell} - VOLUME')
             EOM_Vol = cell # Column for pasting volume data
     elif re.search(r'Net.*Sales', str(forecast_sheet.cell(row=1, column=cell).value)):
         if forecast_sheet.cell(row=2, column=cell).value == reporting_month:
             print(f'{cell} - NS')
             EOM_Sales = cell # Column for pasting sales data
     elif re.search(r'JM', str(forecast_sheet.cell(row=1, column=cell).value)):
         if forecast_sheet.cell(row=2, column=cell).value == reporting_month:
             print(f'{cell} - JM')
             EOM_JM = cell # Column for pasting JM data

for cell in range(3, forecast_sheet.max_row):
    customer_cell = forecast_sheet.cell(row=cell, column=11).value
    prod_ID = forecast_sheet.cell(row=cell, column=14).value
    product_name = forecast_sheet.cell(row=cell, column=15).value

    if customer_cell in sales_data:
        print(f'{customer_cell} data found. Checking Product ID')

        if prod_ID in sales_data[customer_cell]:
            print(f'{customer_cell} matches {prod_ID}. Checking Product Name')

            if product_name in sales_data[customer_cell][prod_ID]:
                print(f'{product_name} matches {prod_ID} and {customer_cell}. Giving value')
                print(f"Sales: {sales_data[customer_cell][prod_ID][product_name]['sales']}")
                forecast_sheet.cell(row=cell, column=19).value = ''  # Cleaning up the MTD DATA
                forecast_sheet.cell(row=cell, column=37).value = ''
                forecast_sheet.cell(row=cell, column=55).value = ''
                forecast_sheet.cell(row=cell, column=EOM_Vol).value = sales_data[customer_cell][prod_ID][product_name][
                    'volume']
                forecast_sheet.cell(row=cell, column=EOM_Sales).value = sales_data[customer_cell][prod_ID][product_name][
                    'sales']
                forecast_sheet.cell(row=cell, column=EOM_JM).value = sales_data[customer_cell][prod_ID][product_name][
                    'profit']

                try:
                    del unmatched_sales[customer_cell][prod_ID]  # delete sales line, because it has been matched

                    if unmatched_sales[customer_cell] == {}:
                        del unmatched_sales[customer_cell]
                        # If no more values/products are left, we delete the whole customer
                        continue
                except KeyError:
                    pyautogui.alert(
                    f"{unmatched_sales[customer_cell][prod_ID]} seems to have a duplicate in the FC.\n"
                    f"Please delete the duplicate row and try again.")
                    sys.exit()

    # WORKS!!! Now we have sales_data as our complete data-set, while unmatched_sales has the non-pasted lines

# From unmatched_data, we need to check whether unmatched_data[customer][prodID] is Delivery'/'Storage fee ID
for cust in sales_data:
    # we will only need the total delivery/storage fee, so we concatinate all of them into a new dic
    for prodID in sales_data[cust]:
        if prodID in ['00001', '00003', '00008']:
            for prodN in sales_data[cust][prodID]:
                storage_delivery_trade['delivery_storage'] += sales_data[cust][prodID][prodN]['sales']

                del unmatched_sales[cust][prodID]  # deleting line, because it has been matched

                if unmatched_sales[cust] == {}:
                    del unmatched_sales[cust]

# We want to know the TOTAL volume/sales/profit from the sales business and paste it in the trade business row
for cust in sales_data:
    if cust in trade_customers:  # trade_customers has been defined at start of script. For if we need to add more
        for prodID in sales_data[cust]:
            for prodN in sales_data[cust][prodID]:  # Only concatinating the int
                storage_delivery_trade['trade_volume'] += sales_data[cust][prodID][prodN]['volume']
                storage_delivery_trade['trade_sales'] += sales_data[cust][prodID][prodN]['sales']
                storage_delivery_trade['trade_profit'] += sales_data[cust][prodID][prodN]['profit']  # see next if

                if sales_data[cust][prodID][prodN]['profit'] != 0:  # we will still copy to xlsx, but will warn user
                    pyautogui.alert(f'{sales_data[cust][prodID]} - Profit should be 0 as trading business,\n'
                                    f'Keeping this line in the output file for checking,\n'
                                    f'Please confirm whether the COGS is OKAY. Profit has been added to JM')
                    del unmatched_sales[cust][prodID][prodN]['sales']
                    del unmatched_sales[cust][prodID][prodN]['volume']
                else:
                    del unmatched_sales[cust][prodID]

                if unmatched_sales[cust] == {}:
                    del unmatched_sales[cust]

for cell in range(1, forecast_sheet.max_row + 1):
    if forecast_sheet.cell(row=cell, column=15).value == 'Trade Business (Nagano/Haruna/Shoei)':
        forecast_sheet.cell(row=cell, column=19).value = ''
        forecast_sheet.cell(row=cell, column=37).value = ''
        forecast_sheet.cell(row=cell, column=55).value = ''
        forecast_sheet.cell(row=cell, column=EOM_Vol).value = storage_delivery_trade['trade_volume']
        print('Pasted trade volume.')
        forecast_sheet.cell(row=cell, column=EOM_Sales).value = storage_delivery_trade['trade_sales'] + \
                                                         storage_delivery_trade[
                                                             'delivery_storage']  # combining for Net sales
        print('Pasted trade sales.')
        forecast_sheet.cell(row=cell, column=EOM_JM).value = storage_delivery_trade['trade_profit'] + \
                                                         storage_delivery_trade['delivery_storage']
        # DeliveryStorage has no Cogs, so all profit
        print('Pasted trade profit.')
        break

# From this section on we will start writing the data to a file.
# Unmatched data and trade business will show in the file for future references

# Step 4. Formatting the excel sheet, so that it can be used for the next reporting month

# Change MTD header to next month, Change YTD header to current reporting month
next_reporting_month = datetime.datetime.strptime(reporting_month, '%b')
next_reporting_month = next_reporting_month.replace(month=next_reporting_month.month + 1)
next_reporting_month = next_reporting_month.strftime('%b')  # Converts the current month + 1, so the month after

MTD_Naming = forecast_sheet.cell(row=2, column=19).value
forecast_sheet.cell(row=2, column=19).value = re.sub('\w\w\w$', next_reporting_month, MTD_Naming) # changing MTD Header

# Change YTD header to current completed reporting month
YTD_Naming = forecast_sheet.cell(row=2, column=17).value
forecast_sheet.cell(row=2, column=17).value = re.sub('\d\d\d\d\w\w\w$', f'{datetime.datetime.now().year}{reporting_month}',
                                                     YTD_Naming)  # Changing YTD Header

# Grab the columns of EOM, change the formula in the cells so that it also grabs this one, do this for all rows
# In value of cell, we it must be '=SUM(T' as starting, if it is e.g. '=SUM(Q', then this is no relevant
greenFill = PatternFill(start_color='0099CC00', end_color='0099CC00', fill_type='solid')  # Filling cell to green

for cell in range(1, forecast_sheet.max_row + 1):
    if not forecast_sheet.cell(row=cell, column=17).value:  # Continues over empty cells
        continue
    if forecast_sheet.cell(row=cell, column=17).value.startswith('=SUM(T'):
        forecast_sheet.cell(row=cell, column=17).value = f'=SUM(T{cell}:{get_column_letter(EOM_Vol)}{cell})'  # '=SUM(T3:Y3)
        forecast_sheet.cell(row=cell, column=33).value = f'=SUM(AL{cell}:{get_column_letter(EOM_Sales)}{cell})'  # sales
        forecast_sheet.cell(row=cell, column=51).value = f'=SUM(BD{cell}:{get_column_letter(EOM_JM)}{cell})'  # sales

        # Colorize the row of the current reporting month
        forecast_sheet.cell(row=cell, column=EOM_Vol).fill = greenFill
        forecast_sheet.cell(row=cell, column=EOM_Sales).fill = greenFill
        forecast_sheet.cell(row=cell, column=EOM_JM).fill = greenFill

# Step 5. Saving files

while True:
    try:
        today = datetime.datetime.now().strftime('%Y-%m-%d %H.%M')  # will add this to end of file name

        confirm = pyautogui.confirm(f'Updating forecast complete and saving in file path {forecast_folder}.\n\n\n'
                                    f'Do you want to save in a different file path?',
                                    buttons=['Save Now', 'Choose Other Path'])
        if confirm == 'Choose Other Path':
            save_path = pyautogui.prompt('Please insert designated file path for saving.')

            if not save_path:
                quit_program = pyautogui.confirm('Quit program?', buttons=['Yes', 'No'])
                if quit_program == 'Yes':
                    sys.exit()
                continue

            save_path = os.path.abspath(save_path)

        else:
            save_path = forecast_folder

        with open(os.path.join(save_path, f'Sales FC {reporting_month} FINAL UNMATCHED.txt'), 'w') as output:
            unmatched_data_output = pprint.pformat(unmatched_sales)
            storage_delivery_trade_output = pprint.pformat(storage_delivery_trade)
            output.write(f'{unmatched_data_output}\n\n TRADE BUSINESS:\n{storage_delivery_trade_output}')

        forecast_wb.save(os.path.join(save_path, f'Sales FC {reporting_month} FINAL.xlsx'))

        forecast_wb.save(os.path.join(save_path, f'Sales FC {next_reporting_month} NEW BOM {today}.xlsx'))

        pyautogui.alert('Saved succesfully!')
        break
    except FileNotFoundError:
        pyautogui.alert('File directory does not exist, please input an existing file directory!')