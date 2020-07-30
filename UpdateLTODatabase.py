import sqlite3
import openpyxl as pyxl
from LTODateConverterFunction import *

def updateLTO(workbook):
    connect = sqlite3.connect("databases\\LTO.db")
    c = connect.cursor()

    c.execute("CREATE TABLE IF NOT EXISTS lto (id INT, name TEXT, cus TEXT, kam TEXT, stage TEXT, "
              "launch DATE, sales INT, vol INT, status TEXT, comment TEXT)")

    global wb, ws
    wb = workbook  # Using the function lto_date_format from our LTO date converter
    ws = wb.active

    # Concatenating each change of the database, will return this
    new_adds = 0
    existing_adds = 0
    closed_adds = 0

    c.execute("SELECT id FROM lto")
    data = c.fetchall()

    global existing_data, uploaded_data
    existing_data = []  # Will append all existing opportunity IDs in order to prevent overwriting existing data
    uploaded_data = []  # Appending all opp IDs to compare it with existing_data. We want to delete rows that are not
    # in the uploaded_data. These rows are overexcess and are closed in the LTO system
    for opp_id in range(len(data)):
        existing_data.append(data[opp_id][0])

    for cell in range(2, ws.max_row + 1):

        id = int(ws.cell(row=cell, column=1).value)
        uploaded_data.append(id)  # Appending only the ID for comparing and removing rows that are closed opportunities
        name = ws.cell(row=cell, column=2).value
        cus = ws.cell(row=cell, column=3).value
        kam = ws.cell(row=cell, column=6).value
        stage = ws.cell(row=cell, column=9).value
        launch = ws.cell(row=cell, column=12).value
        sales = ws.cell(row=cell, column=17).value
        vol = ws.cell(row=cell, column=18).value
        status = "None"
        comment = "None"

        # Testing the opportunity ID for if it already is in the database, will update BUT keep the comment and status
        if id in existing_data:
            existing_adds += 1
            data_exist = [name, cus, kam, stage, launch, sales, vol, id]
            c.execute("UPDATE lto SET name = ?, "
                      "cus = ?, "
                      "kam = ?, "
                      "stage = ?, "
                      "launch = ?, "
                      "sales = ?, "
                      "vol = ? WHERE id = ?", data_exist)
            connect.commit()
        else:
            new_adds += 1
            data_new = [id, name, cus, kam, stage, launch, sales, vol, status, comment]
            c.execute("INSERT INTO lto VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", data_new)
            connect.commit()

    # Comparing the existing_data and uploaded_data. If there is an ID in the database that is not in the uploaded_data
    # We then delete that ID and its row from the database, as it is an old value that is probably closed
    for remove_data in existing_data:
        if remove_data not in uploaded_data:
            closed_adds += 1
            print(f"ID {remove_data} does not exist in uploaded data. This is old data, removing from database...")
            c.execute("DELETE FROM lto WHERE id = ?", (str(remove_data),))
            connect.commit()

    c.execute("SELECT * FROM lto")
    data = c.fetchall()
    print(f"New added: {new_adds}\n"
          f"Existing adjusted: {existing_adds}\n"
          f"Removed old data: {closed_adds}\n\n"
          f"Total in Table: {len(data)}")

    c.close
    connect.close()

    return [new_adds, existing_adds, closed_adds, len(data)]

oksir = pyxl.load_workbook("C:\\Users\\woens\\Desktop\\July Reporting\\test.xlsx")
updateLTO(oksir)
